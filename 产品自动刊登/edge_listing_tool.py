import time
import re
import traceback
import datetime
import openpyxl
from PyQt5.QtCore import QThread, pyqtSignal
from selenium import webdriver
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException


class LocatorParser:
    @staticmethod
    def parse(locator_str: str) -> tuple:
        locator_str = locator_str.strip()
        if not locator_str: return None, None
        if locator_str.startswith('./'): return (By.XPATH, locator_str)
        if locator_str.startswith('//') or locator_str.startswith('(') or locator_str.startswith('.//'): return (
            By.XPATH, locator_str)
        if locator_str.startswith('#') or locator_str.startswith('.'): return (By.CSS_SELECTOR, locator_str)
        attr_match = re.match(r'^([\w-]+)=\"(.*?)\"$', locator_str) or re.match(r"^([\w-]+)='(.*?)'$", locator_str)
        if attr_match: return (By.CSS_SELECTOR, f"[{attr_match.group(1)}='{attr_match.group(2)}']")
        if locator_str.startswith("<span>"):
            text = locator_str.replace("<span>", "").strip()
            return (By.XPATH, f".//span[normalize-space(text())='{text}']")
        return (By.XPATH, f".//*[normalize-space(text())='{locator_str}']")


class RestartSkuException(Exception):
    pass


class ListingWorker(QThread):
    log_signal = pyqtSignal(str, str)
    finished_signal = pyqtSignal()
    error_signal = pyqtSignal(str)
    pause_required_signal = pyqtSignal(str)

    def __init__(self, config_data, is_headless, sku_list=None, excel_path=None):
        super().__init__()
        self.config_data = config_data
        self.is_headless = is_headless
        self.sku_list = sku_list or []
        self.excel_path = excel_path
        self.is_running = True
        self.is_paused = False
        self.driver = None
        self.shop_name = config_data.get('ACCOUNT_NAME', '')
        self.current_site_index = 0
        self.need_restart_current_sku = False

    def stop(self):
        self.is_running = False
        self.is_paused = False
        self.requestInterruption()

    def _log(self, msg, color="black"):
        self.log_signal.emit(msg, color)
        try:
            date_str = datetime.datetime.now().strftime("%Y-%m-%d")
            with open(f"log_{date_str}.txt", "a", encoding="utf-8") as f:
                f.write(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {msg}\n")
        except:
            pass

    def _init_driver(self):
        options = EdgeOptions()
        if self.is_headless:
            options.add_argument("--headless")
            options.add_argument("--disable-gpu")
        else:
            options.add_argument("--start-maximized")
        options.add_argument("--ignore-certificate-errors")
        options.add_argument("--log-level=3")
        return webdriver.Edge(options=options)

    def _parse_config(self):
        parsed = {}
        for mod in self.config_data.get('ELEMENT_CONFIG', []):
            for ele in mod['elements']:
                by, val = LocatorParser.parse(ele['locator'])
                if by:
                    parsed[ele['name']] = {
                        'locator': (by, val),
                        'timeout': ele.get('timeout', 10),
                        'rest': ele.get('rest', 2)
                    }
        return parsed

    def request_manual_pause(self):
        self.is_paused = True
        self._log("⏸️ 用户请求暂停...", "orange")

    def resume_work(self, new_config_data=None):
        if new_config_data:
            self.config_data = new_config_data
            self._log("🔄 配置已更新", "green")
        self.is_paused = False
        self.need_restart_current_sku = True
        self._log("▶️ 准备重试当前SKU流程...", "green")

    def _check_pause(self):
        if not self.is_running: return
        if self.is_paused:
            self.pause_required_signal.emit("流程受阻或手动暂停")
            while self.is_paused and self.is_running:
                time.sleep(1)
            if self.need_restart_current_sku:
                self.need_restart_current_sku = False
                raise RestartSkuException("User resumed from pause")

    def _wait_visible_then_rest(self, driver, name, timeout=None, rest=None, root=None, optional=False):
        self._check_pause()
        ctx = root if root else driver
        cfg = self._parse_config().get(name)
        if not cfg:
            if optional: return None
            self._log(f"❌ 配置缺失: {name}", "red");
            self.is_paused = True;
            self._check_pause();
            return None

        use_timeout = cfg['timeout'] if timeout is None else timeout
        use_rest = cfg['rest'] if rest is None else rest
        self._log(f"🔎 正在定位: [{name}]...", "gray")
        try:
            el = WebDriverWait(ctx, use_timeout).until(EC.visibility_of_element_located(cfg['locator']))
            self._highlight(driver, el, "green")
            log_color = "gray" if optional else "green"
            self._log(f"   ✅ 成功锁定: [{name}]", log_color)
            if use_rest > 0: time.sleep(use_rest)
            return el
        except TimeoutException:
            if optional: return None
            self._log(f"❌ 超时失败: [{name}] 未在 {use_timeout}s 内可见", "red")
            self.is_paused = True;
            self._check_pause()
        except RestartSkuException:
            raise
        except Exception as e:
            if optional: return None
            self._log(f"❌ 异常 [{name}]: {e}", "red");
            self.is_paused = True;
            self._check_pause()
        return None

    def _find(self, driver, name, root=None, optional=False):
        return self._wait_visible_then_rest(driver, name, root=root, optional=optional)

    def _find_in_root(self, root, name, optional=False):
        return self._wait_visible_then_rest(self.driver, name, root=root, optional=optional)

    def _highlight(self, driver, element, color="red"):
        try:
            driver.execute_script(f"arguments[0].style.border='2px solid {color}'", element)
        except:
            pass

    def _safe_click(self, driver, element, name=""):
        if not element: return False
        action_name = name if name else "元素"
        self._log(f"   🖱️ 点击: {action_name}", "black")
        try:
            element.click();
            return True
        except:
            try:
                driver.execute_script("arguments[0].click();", element);
                return True
            except Exception as e:
                self._log(f"❌ 点击失败 [{action_name}]: {e}", "red");
                return False

    def _wait_loading_mask(self, driver, timeout=15):
        try:
            WebDriverWait(driver, timeout).until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, ".el-loading-mask")))
        except:
            pass

    # ==========================================
    # 🌟 复杂 Root 定位逻辑
    # ==========================================
    def _get_complex_root(self, driver, timeout=20):
        self._check_pause()
        self._log("🔎 正在扫描: [动态 Root] (复杂逻辑)...", "gray")
        end_time = time.time() + timeout

        while time.time() < end_time:
            if not self.is_running: return None
            try:
                siblings = driver.find_elements(By.XPATH, "//body/textarea/following-sibling::div")
                for div in siblings:
                    if not self._is_node_active(div): continue
                    try:
                        child = div.find_element(By.XPATH, "./div[1]")
                        if not self._is_node_active(child): continue
                        self._highlight(driver, div, "green")
                        self._log("   ✅ 成功锁定: [动态 Root]", "green")
                        time.sleep(1)
                        return div
                    except:
                        continue
            except Exception as e:
                pass
            time.sleep(1)

        self._log("❌ [动态 Root] 定位超时", "red")
        self.is_paused = True
        self._check_pause()
        return None

    def _is_node_active(self, element):
        try:
            if not element.is_displayed(): return False
            style = element.get_attribute("style") or ""
            if "display: none" in style: return False
            return True
        except:
            return False

    def _get_active_ai_root(self, driver, timeout=0):
        self._check_pause()
        cfg = self._parse_config().get("AI弹窗_Root")
        if not cfg: return None
        end = time.time() + timeout
        while True:
            if not self.is_running: return None
            try:
                all_popups = driver.find_elements(*cfg['locator'])
                if self.current_site_index < len(all_popups):
                    target = all_popups[self.current_site_index]
                    child = target.find_element(By.XPATH, "./div[1]")
                    if "display: none" not in (child.get_attribute("style") or ""):
                        self._highlight(driver, target, "blue");
                        return target
            except:
                pass
            if timeout == 0: break
            if time.time() >= end: break
            time.sleep(0.5)
        return None

    def _get_active_infringement_root(self, driver, timeout=0):
        self._check_pause()
        end = time.time() + timeout
        while True:
            if not self.is_running: return None
            try:
                wrappers = driver.find_elements(By.XPATH, "//div[contains(@class, 'ivu-modal-wrap')]")
                for w in wrappers:
                    if not w.is_displayed(): continue
                    try:
                        content_text = w.get_attribute("innerText")
                        if "侵权词/敏感词/商标词/黑名单" not in content_text: continue
                        confirm_btns = w.find_elements(By.XPATH, ".//button[contains(., '确定')]")
                        has_visible = False
                        for btn in confirm_btns:
                            if btn.is_displayed(): has_visible = True; break
                        if has_visible:
                            modal_content = w.find_element(By.XPATH, ".//div[contains(@class, 'ivu-modal-content')]")
                            self._highlight(driver, modal_content, "orange");
                            return modal_content
                    except:
                        pass
            except:
                pass
            if timeout == 0: break
            if time.time() >= end: break
            time.sleep(0.5)
        return None

    def _force_close_popups(self):
        ai = self._get_active_ai_root(self.driver, timeout=0.5)
        if ai:
            btn = self._find_in_root(ai, "AI弹窗_取消按钮", optional=True)
            if btn: self._safe_click(self.driver, btn, "关闭AI"); time.sleep(0.5)
        inf = self._get_active_infringement_root(self.driver, timeout=0.5)
        if inf:
            try:
                btn = inf.find_element(By.XPATH, ".//button[contains(., '取消')]")
                if btn.is_displayed(): self._safe_click(self.driver, btn, "关闭侵权"); time.sleep(0.5)
            except:
                pass

    # ==========================================
    # 🚀 主流程
    # ==========================================
    def run(self):
        login_retry = 0
        while self.is_running and login_retry < 3:
            try:
                self._log("🚀 启动任务...", "blue")
                self.driver = self._init_driver()
                if not self._step_1_login():
                    raise Exception("登录失败")
                if not self._step_2_nav_to_listing(): raise Exception("导航失败")
                self._step_3_sku_loop()
                self._log("✅ 所有 SKU 处理完毕", "green")
                self.finished_signal.emit()
                break
            except Exception as e:
                self._log(f"❌ 全局异常: {e}", "red")
                traceback.print_exc()
                if self.driver:
                    try:
                        self.driver.quit()
                    except:
                        pass
                login_retry += 1
                time.sleep(3)
        if login_retry >= 3 and self.is_running: self.error_signal.emit("三次启动尝试均失败")

    def _step_1_login(self):
        self._log("--- 步骤1: 登录 ---", "blue")
        try:
            self.driver.get("https://saaserp-pos.yibainetwork.com")
            try:
                self._wait_visible_then_rest(self.driver, '账号输入框', optional=True)
            except:
                pass
            if "login" in self.driver.current_url:
                user_in = self._find(self.driver, '账号输入框')
                user_in.clear()
                user_in.send_keys(self.config_data.get('USERNAME', ''))
                pwd_in = self._find(self.driver, '密码输入框')
                pwd_in.clear()
                pwd_in.send_keys(self.config_data.get('PASSWORD', ''))
                self._safe_click(self.driver, self._find(self.driver, '登录按钮'), "登录按钮")
                confirm_btn = self._wait_visible_then_rest(self.driver, '确认登录按钮', optional=True)
                if confirm_btn:
                    org_in = self._find(self.driver, '组织输入框', optional=True)
                    if org_in:
                        org_in.send_keys(self.config_data.get('ORG_CODE', '156'))
                        time.sleep(0.5)
                        self._safe_click(self.driver, self._find(self.driver, '组织列表项', optional=True))
                    self._safe_click(self.driver, confirm_btn, "确认登录按钮")
            WebDriverWait(self.driver, 20).until(EC.url_contains("home_page"))
            self._log("✅ 登录成功", "green")
            return True
        except RestartSkuException:
            raise
        except Exception as e:
            self._log(f"登录异常: {e}", "red")
            return False

    def _step_2_nav_to_listing(self):
        self._log("--- 步骤2: 切换刊登 ---", "blue")
        try:
            self._check_pause()
            erp_menu = self._find(self.driver, '导航_ERP菜单')
            ActionChains(self.driver).move_to_element(erp_menu).perform()
            nav_btn = self._wait_visible_then_rest(self.driver, '导航_刊登管理')
            if not nav_btn: return False
            handles_before = self.driver.window_handles
            self._safe_click(self.driver, nav_btn, "刊登管理菜单")
            WebDriverWait(self.driver, 10).until(EC.new_window_is_opened(handles_before))
            new_window = [w for w in self.driver.window_handles if w not in handles_before][0]
            self.driver.switch_to.window(new_window)
            self.work_window_handle = new_window
            WebDriverWait(self.driver, 15).until(EC.url_contains("message_center"))
            self._log("✅ 进入 Message Center", "green")
            return True
        except RestartSkuException:
            raise
        except Exception as e:
            self._log(f"导航异常: {e}", "red")
            return False

    def _step_3_sku_loop(self):
        for i, sku in enumerate(self.sku_list):
            if not self.is_running: break
            self._log(f"📦 [进度 {i + 1}/{len(self.sku_list)}] 处理 SKU: {sku}", "blue")
            success = False
            for retry in range(2):
                if not self.is_running: break
                try:
                    self._process_single_sku_flow(sku)
                    self._update_excel_status(sku)
                    success = True
                    break
                except RestartSkuException:
                    self._log(f"🔁 配置已修复，重试...", "blue");
                    self._recover_page_state();
                    continue
                except Exception as e:
                    self._log(f"⚠️ 失败重试: {e}", "orange");
                    self._recover_page_state()
            if not success: self._log(f"❌ SKU {sku} 失败", "red")

    def _recover_page_state(self):
        try:
            self.driver.refresh();
            time.sleep(5);
            self._wait_loading_mask(self.driver)
        except:
            pass

    def _process_single_sku_flow(self, sku):
        self._check_pause()
        try:
            nav = self._find(self.driver, '菜单_刊登管理', optional=True)
            if nav: ActionChains(self.driver).move_to_element(nav).perform(); self._safe_click(self.driver,
                                                                                               self._find(self.driver,
                                                                                                          '菜单_产品列表',
                                                                                                          optional=True))
        except:
            pass
        self._wait_loading_mask(self.driver)
        WebDriverWait(self.driver, 15).until(EC.url_contains("product_list"))
        self._search_sku_logic(sku)
        time.sleep(2)
        try:
            btn_cfg = self._parse_config().get('弹窗_下一步按钮')
            if btn_cfg:
                for btn in self.driver.find_elements(*btn_cfg['locator']):
                    if btn.is_displayed(): self._safe_click(self.driver, btn); self._wait_loading_mask(
                        self.driver); time.sleep(2); break
        except:
            pass

        root = self._get_complex_root(self.driver)
        if not root: raise Exception("无法定位 Root (复杂逻辑失败)")

        body = self._find_in_root(root, '容器_Body')
        if not body: raise Exception("无法定位 Body")
        shop_container = self._find_in_root(body, '容器_店铺区域')
        if not shop_container: raise Exception("无法定位 店铺区域")
        self._select_shop_logic(shop_container)

        if not self._wait_for_site_loading_strict(body): raise Exception("站点加载超时")

        # === [调用核心多站点逻辑] ===
        self._execute_multi_site_logic(body)

        self._log("🏁 退出当前 SKU", "black")
        self._force_close_popups()

        btn_module = self._get_buttons_module(body)
        if btn_module:
            active_span = self._get_active_site_btn_container(btn_module)
            if active_span:
                if self._safe_click(self.driver, self._find_in_root(active_span, "按钮_取消", optional=True), "取消"):
                    time.sleep(1)
                    confirm = self._find(self.driver, "退出确认弹窗_确定按钮", optional=True)
                    if confirm: self._safe_click(self.driver, confirm, "确认退出")
        self._wait_loading_mask(self.driver)
        time.sleep(2)

    def _search_sku_logic(self, sku):
        inp = self._wait_visible_then_rest(self.driver, '搜索_SKU输入框')
        inp.clear();
        inp.send_keys(sku)
        self._safe_click(self.driver, self._find(self.driver, '搜索_查询按钮'), "查询")
        time.sleep(3);
        self._wait_loading_mask(self.driver)
        all_btns = self.driver.find_elements(*self._parse_config().get('列表_刊登按钮')['locator'])
        vis = [b for b in all_btns if b.is_displayed()]
        if len(vis) == 1:
            self._safe_click(self.driver, vis[0], "精细刊登")
        else:
            raise Exception("商品搜索不唯一")

    def _select_shop_logic(self, shop_container):
        shop_in = self._find_in_root(shop_container, '店铺_输入框')
        if not shop_in: raise Exception("找不到店铺输入框")
        shop_in.click();
        shop_in.clear();
        shop_in.send_keys(self.shop_name);
        time.sleep(1)
        try:
            self.driver.find_element(By.XPATH, self._parse_config().get('店铺_下拉列表项')['locator'][
                1] + f"[normalize-space(text())='{self.shop_name}']").click()
        except:
            shop_in.send_keys(Keys.ENTER)

    # ==========================================
    # 🌟 [修改] 使用全局提交按钮作为加载标识
    # ==========================================
    def _wait_for_site_loading_strict(self, body_root):
        self._log("⏳ 等待站点加载 (检测全局按钮)...", "blue")
        target_btn_name = '按钮_提交所有'
        btn_cfg = self._parse_config().get(target_btn_name)

        # 1. 尝试等待按钮出现
        if btn_cfg:
            end = time.time() + btn_cfg.get('timeout', 30)
            while time.time() < end:
                self._check_pause()
                try:
                    btns = body_root.find_elements(*btn_cfg['locator'])
                    # 只要按钮存在且显示，即视为加载完成
                    if len([b for b in btns if b.is_displayed()]) >= 1:
                        return True
                except:
                    pass
                time.sleep(1)

        # 2. [新增] 按钮超时未现身，进行兜底检查：是否所有站点都已推送？
        self._log("⚠️ 全局按钮未出现，检查是否【全站已推送】...", "orange")
        try:
            # 获取 Tab 区域
            tabs_container = self._find_in_root(body_root, '容器_Tabs区域', optional=True)
            if tabs_container:
                items = tabs_container.find_elements(By.CSS_SELECTOR, "span.item")
                if items:
                    all_pushed = True
                    # 检查每一个 Tab 的文字
                    for item in items:
                        txt = item.get_attribute("textContent")
                        # 如果有一个没包含 "已推送"，说明页面还没加载完或者真的出错了
                        if "已推送" not in txt:
                            all_pushed = False
                            break

                    if all_pushed:
                        self._log("✅ 检测到所有站点均为 [已推送] 状态，跳过处理", "green")
                        # 返回 True 让流程继续，后续的 _execute_multi_site_logic 会自动跳过每个站点
                        return True
                    else:
                        self._log("❌ 检测到仍有未推送站点，但按钮缺失", "red")
        except Exception as e:
            self._log(f"❌ 兜底检查异常: {e}", "red")

        return False

    def _get_copy_module(self, body_root):
        try:
            main = self._find_in_root(body_root, '容器_Main')
            layout = self._find_in_root(main, '容器_布局Wrapper')
            site_divs = layout.find_elements(By.XPATH, "./div")
            active = next((s for s in site_divs if s.is_displayed()), None)
            if not active: return None
            return self._find_in_root(self._find_in_root(active, '容器_站点模块Wrapper'), '容器_文案模块')
        except:
            return None

    def _get_buttons_module(self, body_root):
        try:
            return self._find_in_root(self._find_in_root(body_root, '容器_Main'), '容器_按钮模块')
        except:
            return None

    def _get_active_site_btn_container(self, btn_module):
        if not btn_module: return None
        try:
            spans = btn_module.find_elements(By.XPATH, "./span[contains(@class, 'f-btn')]")
            for s in spans:
                if s.is_displayed():
                    self._highlight(self.driver, s, "blue");
                    return s
        except:
            pass
        return None

    def _get_global_submit_btn(self, btn_module, timeout=5):
        if not btn_module: return None
        end = time.time() + timeout
        while time.time() < end:
            try:
                btns = btn_module.find_elements(By.XPATH, "./button")
                if btns:
                    target = btns[-1]
                    if target.is_displayed(): self._highlight(self.driver, target, "purple"); return target
            except:
                pass
            time.sleep(0.5)
        return None

    def _verify_function_buttons(self, body_root):
        self._log("   🧐 验证功能按钮...", "black")
        btn_module = self._get_buttons_module(body_root)
        if not self._get_global_submit_btn(btn_module, timeout=5):
            self._log("❌ 缺失 全局提交按钮", "red");
            self.is_paused = True;
            self._check_pause();
            return
        active_span = self._get_active_site_btn_container(btn_module)
        if not active_span:
            self._log("❌ 缺失 站点按钮容器", "red");
            self.is_paused = True;
            self._check_pause();
            return
        for name in ["按钮_取消", "按钮_同步", "按钮_翻译", "按钮_保存当前", "按钮_保存所有", "按钮_提交当前"]:
            if not self._find_in_root(active_span, name, optional=True):
                self._log(f"❌ 缺失按钮: {name}", "red");
                self.is_paused = True;
                self._check_pause()
        self._log("   ✅ 按钮校验通过", "green")

    def _execute_multi_site_logic(self, body_root):
        tabs = self._find_in_root(body_root, '容器_Tabs区域')
        items = tabs.find_elements(By.CSS_SELECTOR, "span.item")
        total = len(items)
        self._log(f"📋 共检测到 {total} 个站点", "blue")

        is_first_operation = True

        for i in range(total):
            if not self.is_running: break
            self.current_site_index = i

            tabs = self._find_in_root(body_root, '容器_Tabs区域')
            items = tabs.find_elements(By.CSS_SELECTOR, "span.item")
            if i >= len(items): break
            item = items[i]

            site_name = item.get_attribute("textContent").strip()

            if "已推送" in site_name:
                self._log(f"👉 站点 {i + 1} ({site_name}) 已推送 -> 跳过", "gray")
                continue

            self._log(f"👉 处理站点 {i + 1} ({site_name})", "blue")
            try:
                self._safe_click(self.driver, item, f"切换站点-{site_name}")
            except Exception as e:
                # 最后的保底：纯 JS 强点
                self.driver.execute_script("arguments[0].click();", item)
            self._wait_loading_mask(self.driver)
            time.sleep(3)

            self._verify_function_buttons(body_root)

            if is_first_operation:
                self._log("🌟 [首站逻辑] 开始执行全套流程...", "blue")
                self._flow_ai_generation(body_root)
                self._flow_infringement_check(body_root)
                self._click_and_wait(body_root, "按钮_保存当前")
                self._click_and_wait(body_root, "按钮_同步", wait_time=10)
                self._click_and_wait(body_root, "按钮_翻译", wait_time=10)
                self._click_and_wait(body_root, "按钮_保存所有")
                self._submit_current_and_handle_errors(self.driver, body_root)

                is_first_operation = False
            else:
                self._log("⏩ [后续站点] 执行简化流程...", "blue")
                self._flow_infringement_check(body_root, wait_time=5)
                self._submit_current_and_handle_errors(self.driver, body_root)

    def _click_and_wait(self, body_root, btn_name, wait_time=2):
        btn_module = self._get_buttons_module(body_root)
        active_span = self._get_active_site_btn_container(btn_module)
        btn = self._find_in_root(active_span, btn_name)
        if self._safe_click(self.driver, btn, btn_name):
            self._wait_loading_mask(self.driver)
            time.sleep(wait_time)

    def _submit_current_and_handle_errors(self, driver, body_root):
        btn_module = self._get_buttons_module(body_root)
        active_span = self._get_active_site_btn_container(btn_module)
        submit_btn = self._find_in_root(active_span, "按钮_提交当前")

        if not submit_btn:
            self._log("❌ 无法找到提交按钮", "red")
            return

        # 最多重试 3 次
        for attempt in range(1, 4):
            if not self.is_running: break

            self._log(f"🚀 提交当前页 (第 {attempt} 次)...", "blue")

            # 1. 点击提交
            if not self._safe_click(driver, submit_btn, "提交当前"): return

            # 等待潜在的弹窗加载
            self._wait_loading_mask(driver)
            time.sleep(2)

            # 2. 检测是否有 [提示弹窗] (核心逻辑)
            prompt_root = self._get_prompt_popup(driver, timeout=3)

            if prompt_root:
                # === A. 获取错误信息 ===
                error_text = self._extract_prompt_text(prompt_root)
                self._log(f"   ⚠️ 捕获提示信息: {error_text}", "orange")

                # === B. 关闭弹窗 (必须先关闭才能操作页面其他元素) ===
                self._close_prompt_popup(driver, prompt_root)

                # === C. 分析错误并处理 ===
                if "必填项" in error_text:
                    self._log("   🔧 识别到必填项缺失，尝试自动填充...", "blue")
                    self._fill_mandatory_fields()
                    # 处理完后，进入下一次循环(再次提交)
                    continue

                elif "侵权" in error_text or "敏感" in error_text:
                    self._log("   🛡️ 识别到侵权/敏感词，尝试处理...", "blue")
                    # 这里调用之前的侵权检测逻辑，或者简单的确认点击
                    # 如果弹窗只是提示，可能需要重新去点一下"检测侵权"或者手动修
                    # 假设这里调用之前的暴力确认逻辑：
                    self._flow_infringement_check(body_root, wait_time=2)
                    continue

                else:
                    self._log("   ⛔ 遇到无法自动修复的错误，跳过当前站点", "red")
                    break  # 跳出重试，放弃该站点

            # 3. 检测是否有 [侵权确认弹窗] (有些时候不弹提示，直接弹侵权确认)
            # 这种情况通常不需要关闭，直接点确定即可
            inf_root = self._get_active_infringement_root(driver, timeout=1)
            if inf_root:
                self._log(f"   ⚠️ 提交触发侵权确认弹窗 ({attempt})", "orange")
                try:
                    confirm = inf_root.find_element(By.XPATH, ".//button[contains(., '确定')]")
                    self.driver.execute_script("arguments[0].click();", confirm)
                    time.sleep(1)
                    # 侵权点完确定后，可能需要再次提交，也可能直接就过了
                    # 这里选择 continue 再提交一次保险，或者视具体业务而定
                    # 如果系统逻辑是点确定就自动提交了，这里 break 也可以
                    self._wait_loading_mask(driver)
                    continue
                except:
                    pass

            # 4. 如果没有弹窗，检查是否还有 Inline 错误 (页面上的红字)
            # (保留之前的简单检查)
            try:
                errs = driver.find_elements(By.XPATH, "//div[contains(@class, 'ivu-notice') and contains(., '必填')]")
                visible_errs = [e for e in errs if e.is_displayed()]
                if visible_errs:
                    self._log("   ⚠️ 发现页面悬浮报错，填充必填项...", "orange")
                    self._fill_mandatory_fields()
                    continue
            except:
                pass

            # 5. 如果执行到这里，说明没有捕获到任何错误弹窗，视为成功
            self._log("   ✅ 提交动作完成 (未捕获阻断性错误)", "green")
            break

    # --- 辅助: 获取提示弹窗 ---
    def _get_prompt_popup(self, driver, timeout=2):
        cfg = self._parse_config().get("提示弹窗_Root")
        if not cfg: return None
        end = time.time() + timeout
        while time.time() < end:
            try:
                # 查找所有符合条件的弹窗
                popups = driver.find_elements(*cfg['locator'])
                for p in popups:
                    if p.is_displayed():
                        # 双重检查：确保里面有 "提示" 两个字 (防止定位到其他弹窗)
                        if "提示" in p.get_attribute("innerText"):
                            self._highlight(driver, p, "red")
                            return p
            except:
                pass
            time.sleep(0.5)
        return None

    # --- 辅助: 提取弹窗文本 ---
    def _extract_prompt_text(self, prompt_root):
        text = ""
        try:
            cfg = self._parse_config().get("提示弹窗_错误信息")
            if cfg:
                # 修正路径为相对路径
                by, val = cfg['locator']
                if by == By.XPATH and val.startswith("//"): val = "." + val

                spans = prompt_root.find_elements(by, val)
                text = " ".join([s.text.strip() for s in spans])
        except Exception as e:
            text = str(e)
        return text

    # --- 辅助: 关闭提示弹窗 (含兜底方案) ---
    def _close_prompt_popup(self, driver, prompt_root):
        self._log("   ❌ 关闭提示弹窗...", "gray")
        closed = False

        # 方案 1: 点击右上角 X
        try:
            cfg = self._parse_config().get("提示弹窗_关闭图标")
            if cfg:
                by, val = cfg['locator']
                if by == By.XPATH and val.startswith("//"): val = "." + val
                close_btn = prompt_root.find_element(by, val)
                self.driver.execute_script("arguments[0].click();", close_btn)
                time.sleep(1)
                if not prompt_root.is_displayed(): closed = True
        except:
            pass

        if closed: return

        # 方案 2: 点击页面空白处 (1.选择平台)
        self._log("   ⚠️ 弹窗未关闭，尝试点击页面背景...", "orange")
        try:
            target_cfg = self._parse_config().get("页面_空白点击目标")
            if target_cfg:
                target = driver.find_element(*target_cfg['locator'])
                # 这里使用 ActionChains 点击，模拟鼠标真实行为，通常能触发 mask 关闭
                ActionChains(driver).move_to_element(target).click().perform()
                time.sleep(1)
        except Exception as e:
            self._log(f"   ❌ 背景点击失败: {e}", "red")

    def _update_excel_status(self, sku):
        if not self.excel_path: return
        self._log(f"📝 更新 Excel 状态: {sku}", "gray")
        try:
            tabs = self.driver.find_elements(By.CSS_SELECTOR, "div.mult-header-h span.item")
            status_list = []
            for t in tabs:
                txt = t.get_attribute("textContent").strip()
                status_list.append(txt)
            result_str = " | ".join(status_list)

            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active

            target_row = None
            for row in ws.iter_rows(min_row=2, max_col=1):
                if str(row[0].value).strip() == str(sku).strip():
                    target_row = row[0].row
                    break

            if target_row:
                ws.cell(row=target_row, column=5, value=result_str)
                wb.save(self.excel_path)
                self._log("   ✅ Excel 保存成功", "green")
            else:
                self._log("   ⚠️ Excel 中未找到对应 SKU 行", "orange")

        except Exception as e:
            self._log(f"   ❌ Excel 回写失败: {e}", "red")

    def _fill_mandatory_fields(self):
        self._log("   🔧 [智能填充] 扫描报错必填项...", "blue")
        filled_count = 0

        # 1. 核心定位策略：寻找所有显示出来的 "带*号为必填项" 提示
        # 只有显示了这句话，才说明这个地方校验没过，需要处理
        error_tips = self.driver.find_elements(By.XPATH,
                                               "//div[@class='ivu-form-item-error-tip' and contains(text(), '带*号为必填项')]")

        # 过滤出可见的提示 (因为有些可能是隐藏的)
        visible_tips = [tip for tip in error_tips if tip.is_displayed()]

        if not visible_tips:
            self._log("   ⚠️ 未找到可见的必填报错提示", "gray")
            return

        for tip in visible_tips:
            try:
                # 2. 回溯到父级容器 (ivu-form-item-content)
                # 结构: <div class="ivu-form-item-content"> ... <div class="error-tip">...</div> </div>
                container = tip.find_element(By.XPATH, "./..")

                # === 情况 A: 下拉选择框 (Select) ===
                # 特征: 容器内有 .ivu-select
                selects = container.find_elements(By.CSS_SELECTOR, ".ivu-select")
                if selects:
                    self._handle_ivu_select(container, selects[0])
                    filled_count += 1
                    continue

                # === 情况 B: 文本输入框 (Input) ===
                # 特征: 容器内有 .ivu-input-wrapper 或直接有 input
                inputs = container.find_elements(By.CSS_SELECTOR, "input.ivu-input")
                if inputs:
                    self._handle_ivu_input(inputs[0])
                    filled_count += 1
                    continue

                # === 情况 C: 多行文本 (Textarea) ===
                textareas = container.find_elements(By.TAG_NAME, "textarea")
                if textareas:
                    self._handle_ivu_input(textareas[0], is_textarea=True)
                    filled_count += 1
                    continue

            except Exception as e:
                self._log(f"   ❌ 处理某必填项失败: {e}", "red")

        if filled_count > 0:
            self._log(f"   ✅ 已修复 {filled_count} 个必填项", "green")
            # 稍微等待一下，让前端验证逻辑跑完 (Tab 触发后的验证)
            time.sleep(1)

    # --- 辅助: 处理 Input/Textarea ---
    def _handle_ivu_input(self, element, is_textarea=False):
        try:
            self._highlight(self.driver, element, "orange")
            # 清空并输入
            element.clear()
            if is_textarea:
                element.send_keys("Default Description Content for Validation.")
            else:
                element.send_keys("1")

            # 关键：发送 TAB 键触发 blur 事件，让 "带*号为必填项" 消失
            element.send_keys(Keys.TAB)
            time.sleep(0.2)
        except Exception as e:
            pass

    # --- 辅助: 处理 Select 下拉 ---
    def _handle_ivu_select(self, container, select_div):
        try:
            # 1. 点击触发下拉
            selection_box = select_div.find_element(By.CSS_SELECTOR, ".ivu-select-selection")
            self._highlight(self.driver, selection_box, "orange")
            self.driver.execute_script("arguments[0].click();", selection_box)
            time.sleep(0.5)

            # 2. 在 container 内部或全局寻找对应的 dropdown
            # iView 的 dropdown 有时会渲染在 body 根节点，而不是 container 内部
            # 但你提供的 HTML 显示 dropdown 就在 .ivu-select 内部 (或者兄弟节点)
            # 我们先尝试找 select_div 内部的 dropdown

            # 策略：点击后，寻找可见的 li.ivu-select-item
            # 因为下拉框打开后，li 应该是可见的
            # 为了防止点到别的下拉框的选项，我们要尽量找"最近"的

            # 尝试方法 1: 在当前组件结构内找
            items = select_div.find_elements(By.CSS_SELECTOR, ".ivu-select-dropdown-list li.ivu-select-item")
            visible_items = [i for i in items if i.is_displayed()]  # 只要可见的

            # 尝试方法 2: 如果组件是 transfer-dom (渲染在body)，则在全局找所有可见的 li
            if not visible_items:
                # 找全局所有可见的下拉项，通常最后一个就是刚刚点开的那个
                all_items = self.driver.find_elements(By.CSS_SELECTOR, "li.ivu-select-item")
                visible_items = [i for i in all_items if i.is_displayed()]

            if visible_items:
                target_item = visible_items[0]  # 选第一个
                self.driver.execute_script("arguments[0].click();", target_item)

                # 3. 按 Tab 确保触发验证 (焦点回到 selection_box 再按 Tab)
                try:
                    ActionChains(self.driver).move_to_element(selection_box).click().send_keys(Keys.TAB).perform()
                except:
                    pass
            else:
                self._log("   ⚠️ 下拉框无选项", "gray")
                # 兜底：如果没有选项，尝试按一下 ESC 或 TAB 关闭下拉
                ActionChains(self.driver).send_keys(Keys.TAB).perform()

        except Exception as e:
            self._log(f"   ❌ 下拉选择异常: {e}", "red")

    # ==========================================
    # 🌟 [修正] 强制 DOM 交互模式 (无视滚动和遮挡)
    # ==========================================
    def _flow_ai_generation(self, body_root):
        text_source = self.config_data.get('TEXT_SOURCE', '网页AI生成')
        if text_source == '跳过文案': return

        self._log("🤖 准备执行 AI 文案生成 (强制模式)...", "black")
        copy_mod = self._get_copy_module(body_root)
        if not copy_mod:
            self._log("❌ 未找到文案模块", "red")
            return

        # 1. 点击打开 AI 弹窗 (这个按钮通常在视口内，可以用常规方法，也可以强点)
        ai_btn = self._find_in_root(copy_mod, "文案_AI按钮", optional=True)
        if not self._safe_click(self.driver, ai_btn, "AI按钮"): return

        self._log("   ⏳ 等待 AI 弹窗加载...", "black")

        # 2. 定位 AI 弹窗容器
        ai_root = self._get_active_ai_root(self.driver, timeout=8)
        if not ai_root:
            self._log("❌ 未捕获到 AI 弹窗 (超时)", "red")
            return

        self._highlight(self.driver, ai_root, "green")
        self._log("✅ 检测到 AI 弹窗", "green")

        # 获取配置
        gen_cfg = self._parse_config().get("AI弹窗_生成按钮")
        app_cfg = self._parse_config().get("AI弹窗_应用按钮")

        # 3. 循环生成检查
        for attempt in range(1, 4):
            if not self.is_running: return

            # [强取] 检查标题长度 (不滚动，不检查可见性)
            current_len = self._check_title_len(ai_root)
            if current_len > 20:
                self._log(f"   ✨ 标题已生成 (长度:{current_len})", "blue")
                break

            # [强点] 生成按钮
            if gen_cfg:
                self._log(f"   👉 [强制点击] 生成文案 ({attempt})", "blue")
                try:
                    # 直接在 root 下找 DOM 元素，不管是否可见
                    btns = ai_root.find_elements(*gen_cfg['locator'])
                    if btns:
                        # 只要存在就强点 (JS Click 不需要元素在视口内)
                        self.driver.execute_script("arguments[0].click();", btns[-1])
                        time.sleep(5)  # 给 AI 思考时间
                    else:
                        self._log("   ⚠️ DOM中未找到生成按钮", "orange")
                except Exception as e:
                    self._log(f"   ❌ 点击异常: {e}", "red")

            # 等待结果刷新
            time.sleep(2)
            if self._check_title_len(ai_root) > 20:
                self._log("   ✅ 文案生成成功！", "green")
                break

        # 4. [强点] 应用按钮
        if app_cfg:
            try:
                # 同样直接找 DOM，强点
                app_btns = ai_root.find_elements(*app_cfg['locator'])
                if app_btns:
                    self._log("   👉 [强制点击] 应用文案", "black")
                    self.driver.execute_script("arguments[0].click();", app_btns[-1])
                    time.sleep(1)
                else:
                    self._log("⚠️ DOM中未找到应用按钮，尝试关闭", "orange")
                    self._force_close_popups()
            except:
                self._force_close_popups()
        else:
            self._force_close_popups()

        # [修改] 强制获取标题长度 (只查 DOM，不查 Visible)
        def _check_title_len(self, root):
            try:
                cfg = self._parse_config().get("AI弹窗_标题输入框")
                if not cfg: return 0
                # 使用 find_elements 避免报错，且不等待 visible
                inps = root.find_elements(*cfg['locator'])
                if inps:
                    # 获取 value 属性不需要元素可见
                    val = inps[0].get_attribute("value")
                    return len(val) if val else 0
            except:
                pass
            return 0

    def _check_title_len(self, root):
        try:
            inp = self._find_in_root(root, "AI弹窗_标题输入框", optional=True)
            if inp: return len(inp.get_attribute("value"))
        except:
            pass
        return 0

    def _flow_infringement_check(self, body_root, wait_time=10):
        self._log("🛡️ 侵权检测...", "black")
        copy_mod = self._get_copy_module(body_root)
        if not copy_mod: return
        chk_btn = self._find_in_root(copy_mod, "文案_侵权检测按钮", optional=True)
        if not self._safe_click(self.driver, chk_btn, "侵权检测"): return
        time.sleep(wait_time)
        inf_root = self._get_active_infringement_root(self.driver, timeout=5)
        if inf_root:
            self._log("   🚨 发现侵权弹窗", "orange")
            confirmed = False
            for i in range(10):
                try:
                    confirm_btn = inf_root.find_element(By.XPATH, ".//button[contains(., '确定')]")
                    if confirm_btn.is_displayed():
                        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'nearest'});",
                                                   confirm_btn)
                        time.sleep(0.5)
                        self.driver.execute_script("arguments[0].click();", confirm_btn)
                        self._log("   ✅ 已点击确认按钮", "green")
                        confirmed = True
                        break
                except:
                    pass
                time.sleep(1)
            if not confirmed: self._log("   ❌ 未找到确认按钮", "red")
        else:
            self._log("   ✅ 无侵权", "green")

    def _flow_sync_trans(self, body_root):
        self._log("🔄 同步与翻译...", "black")
        btn_module = self._get_buttons_module(body_root)
        active_span = self._get_active_site_btn_container(btn_module)

        sync_btn = self._find_in_root(active_span, "按钮_同步", optional=True)
        if sync_btn:
            self._safe_click(self.driver, sync_btn, "同步")
            time.sleep(10)
        else:
            self._log("   ⚠️ 未找到同步", "gray")

        trans_btn = self._find_in_root(active_span, "按钮_翻译", optional=True)
        if trans_btn:
            self._safe_click(self.driver, trans_btn, "翻译")
            time.sleep(10)
        else:
            self._log("   ⚠️ 未找到翻译", "gray")