import sys
import os
import time
import re
import openpyxl
import requests
import xlsxwriter
from datetime import datetime
from io import BytesIO
from PIL import Image as PilImage

from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QTabWidget, QFormLayout, QScrollArea, QComboBox,
                             QLineEdit, QPushButton, QLabel, QDialog, QGroupBox,
                             QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox, QTextEdit, QFileDialog,
                             QCheckBox, QGridLayout)
from PyQt5.QtCore import Qt, QSettings, QSize, QRect
from PyQt5.QtGui import QFont, QTextCharFormat, QTextCursor, QPixmap, QIcon, QPainter, QPen, QBrush, QColor

from config_manager import config_manager
from edge_automation_tool import LocatorParser

try:
    from selenium import webdriver
    from selenium.webdriver.edge.service import Service
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.action_chains import ActionChains
    from selenium.common.exceptions import TimeoutException, NoSuchElementException
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.common.by import By
    from selenium.webdriver.edge.options import Options as EdgeOptions

    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False


# --- 【模式二专用】图片筛选弹窗 ---
class ImageSelectorDialog(QDialog):
    def __init__(self, all_results, source_file_path, image_session, parent=None):
        super().__init__(parent)
        self.all_results = all_results
        self.source_file_path = source_file_path
        self.image_session = image_session
        self.setWindowTitle("图片筛选与导出 (模式二)")
        self.setGeometry(100, 100, 1200, 800)

        self.layout = QVBoxLayout(self)
        self.layout.addWidget(QLabel("请点击图片进行选择 (红框数字代表导出顺序，再次点击取消，最多 9 张):"))

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        container = QWidget()
        self.container_layout = QVBoxLayout(container)

        self.sku_widgets = []
        for res in all_results:
            sku_box = SKUResultWidget(res, self.image_session)
            self.container_layout.addWidget(sku_box)
            self.sku_widgets.append(sku_box)

        container.setLayout(self.container_layout)
        scroll.setWidget(container)
        self.layout.addWidget(scroll)

        btn_layout = QHBoxLayout()
        export_btn = QPushButton("导出选中结果到 Excel (链接版)")
        export_btn.clicked.connect(self.export_data)
        export_btn.setMinimumHeight(40)
        btn_layout.addWidget(export_btn)

        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.reject)
        btn_layout.addWidget(close_btn)

        self.layout.addLayout(btn_layout)

    def export_data(self):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_dir = os.path.dirname(self.source_file_path) if self.source_file_path else os.getcwd()
        filename = f"筛选导出结果_{timestamp}.xlsx"
        save_path = os.path.join(base_dir, filename)

        workbook = xlsxwriter.Workbook(save_path)
        worksheet = workbook.add_worksheet()

        text_keys = []
        for res in self.all_results:
            for k in res.get('text_info', {}).keys():
                if k not in text_keys:
                    text_keys.append(k)

        img_keys = [f"图片链接_{i + 1}" for i in range(9)]

        headers = ["SKU"] + text_keys + img_keys + ["ERROR"]

        header_format = workbook.add_format({'bold': True, 'bg_color': '#D7E4BC', 'border': 1})
        center_format = workbook.add_format({'valign': 'vcenter', 'border': 1})
        wrap_format = workbook.add_format({'valign': 'vcenter', 'border': 1, 'text_wrap': True})

        for col_idx, header in enumerate(headers):
            worksheet.write(0, col_idx, header, header_format)
            if "图片" in header:
                worksheet.set_column(col_idx, col_idx, 40)  # 链接列宽一点
            elif "描述" in header:
                worksheet.set_column(col_idx, col_idx, 40)
            else:
                worksheet.set_column(col_idx, col_idx, 20)

        current_row = 1

        for i, sku_widget in enumerate(self.sku_widgets):
            original_data = self.all_results[i]
            selected_images = sku_widget.get_selected_images()

            worksheet.write(current_row, 0, original_data['SKU'], center_format)

            col_offset = 1
            for key in text_keys:
                val = original_data.get('text_info', {}).get(key, "")
                fmt = wrap_format if key == "描述" else center_format
                worksheet.write(current_row, col_offset, val, fmt)
                col_offset += 1

            # 写入图片链接
            img_start_col = col_offset
            for j in range(9):
                if j < len(selected_images):
                    worksheet.write(current_row, img_start_col + j, selected_images[j], center_format)
                else:
                    worksheet.write(current_row, img_start_col + j, "", center_format)

            error_col = img_start_col + 9
            err_msg = f"{original_data['ERROR']} - {original_data.get('Details', '')}" if 'ERROR' in original_data else ""
            worksheet.write(current_row, error_col, err_msg, center_format)

            # 恢复普通行高
            worksheet.set_row(current_row, 20)
            current_row += 1

        try:
            workbook.close()
            QMessageBox.information(self, "导出成功", f"文件已保存至:\n{save_path}")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"保存文件时出错: {e}")


# --- 【新增类】可点击的图片 Label ---
class ClickableImageLabel(QLabel):
    def __init__(self, img_url, parent_widget):
        super().__init__()
        self.img_url = img_url
        self.parent_widget = parent_widget
        self.selection_number = 0

        self.setFixedSize(100, 100)
        self.setScaledContents(True)
        self.setCursor(Qt.PointingHandCursor)

        self.default_style = "border: 1px solid #ccc; background-color: #f0f0f0;"
        self.setStyleSheet(self.default_style)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.parent_widget.handle_image_click(self.img_url)

    def set_selection_number(self, number):
        self.selection_number = number
        self.repaint()

    def paintEvent(self, event):
        super().paintEvent(event)
        if self.selection_number > 0:
            painter = QPainter(self)
            painter.setRenderHint(QPainter.Antialiasing)
            pen = QPen(Qt.red)
            pen.setWidth(4)
            painter.setPen(pen)
            painter.drawRect(self.rect())
            badge_size = 24
            badge_rect = QRect(self.width() - badge_size, 0, badge_size, badge_size)
            brush = QBrush(Qt.red)
            painter.setBrush(brush)
            painter.setPen(Qt.NoPen)
            painter.drawRect(badge_rect)
            painter.setPen(Qt.white)
            font = QFont("Arial", 12, QFont.Bold)
            painter.setFont(font)
            painter.drawText(badge_rect, Qt.AlignCenter, str(self.selection_number))


# --- 【模式二专用】单个 SKU 展示组件 ---
class SKUResultWidget(QGroupBox):
    def __init__(self, sku_data, image_session, parent=None):
        super().__init__(parent)
        self.sku_data = sku_data
        self.image_session = image_session
        self.setTitle(f"SKU: {sku_data['SKU']}")
        self.selected_images_list = []
        self.image_label_map = {}
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        if 'ERROR' in self.sku_data:
            error_label = QLabel(f"❌ 错误: {self.sku_data['ERROR']}")
            error_label.setStyleSheet("color: red;")
            layout.addWidget(error_label)

        # 【核心修改】移除了 info_text 显示区域，只保留图片选择

        images = self.sku_data.get('all_images', [])
        if images:
            image_area = QWidget()
            grid = QGridLayout(image_area)

            for i, img_url in enumerate(images):
                img_label = ClickableImageLabel(img_url, self)
                img_label.setToolTip(f"{img_url}")
                self.image_label_map[img_url] = img_label

                # 尝试加载图片用于预览
                if i < 20:
                    try:
                        req = self.image_session.get(img_url, timeout=2)
                        if req.status_code == 200:
                            pixmap = QPixmap()
                            pixmap.loadFromData(req.content)
                            img_label.setPixmap(pixmap)
                        else:
                            img_label.setText("加载失败")
                    except:
                        img_label.setText("Error")
                else:
                    img_label.setText(f"图 {i + 1}")

                grid.addWidget(img_label, i // 8, i % 8)

            layout.addWidget(image_area)
        else:
            layout.addWidget(QLabel("未找到图片"))

        self.setLayout(layout)

    def handle_image_click(self, img_url):
        if img_url in self.selected_images_list:
            self.selected_images_list.remove(img_url)
        else:
            if len(self.selected_images_list) < 9:
                self.selected_images_list.append(img_url)
            else:
                print("最多只能选择 9 张图片")
        self.refresh_labels_state()

    def refresh_labels_state(self):
        for index, url in enumerate(self.selected_images_list):
            label = self.image_label_map.get(url)
            if label:
                label.set_selection_number(index + 1)
        for url, label in self.image_label_map.items():
            if url not in self.selected_images_list:
                label.set_selection_number(0)

    def get_selected_images(self):
        return self.selected_images_list


class AutomationToolUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Edge 自动化配置工具")
        self.setGeometry(100, 100, 950, 800)

        self.config_settings = QSettings('MyCompany', 'EdgeAutoTool')

        self.all_accounts = []
        self.element_config = []
        self.element_widgets = {}
        self.runtime_selected_account_name = ''
        self.sku_file_path = ''
        self.runtime_headless = False
        self.image_session = requests.Session()

        self.load_config()

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)

        self.main_layout = QVBoxLayout(self.central_widget)
        self.tab_widget = QTabWidget()
        self.main_layout.addWidget(self.tab_widget)

        self.create_operation_page()
        self.create_config_page()

    def _unify_element_config(self, code_structure, json_data):
        json_map = {}
        if json_data:
            for module_item in json_data:
                for element in module_item.get("elements", []):
                    json_map[element["name"]] = element
        unified_config = code_structure.copy()
        for module_item in unified_config:
            for element in module_item["elements"]:
                name = element["name"]
                if name in json_map:
                    element["locator"] = json_map[name].get("locator", "")
                    element["position"] = json_map[name].get("position", "当前元素")
                    element["index"] = json_map[name].get("index", "1")
                else:
                    if "position" not in element: element["position"] = "当前元素"
                    if "index" not in element: element["index"] = "1"
        return unified_config

    def load_config(self):
        config = config_manager.load_config()
        default_config = config_manager.default_config
        self.all_accounts = config.get("ACCOUNTS", [])
        self.element_config = self._unify_element_config(
            default_config.get("ELEMENT_CONFIG", []),
            config.get("ELEMENT_CONFIG_FROM_FILE", [])
        )
        self.runtime_url = self.config_settings.value('url', config.get("LOGIN_URL"))
        self.runtime_org_code = self.config_settings.value('org_code', config.get("ORG_CODE"))
        self.sku_file_path = self.config_settings.value('sku_file_path', config.get("SKU_FILE_PATH"))
        self.runtime_start_point = self.config_settings.value('start_point', '完整流程 (从登录开始)')
        self.runtime_headless = self.config_settings.value('headless', 'false') == 'true'
        self.runtime_run_mode = self.config_settings.value('run_mode', '模式一：自动导出 (按分类)')

        first_account_name = self.all_accounts[0]['name'] if self.all_accounts else ''
        self.runtime_selected_account_name = self.config_settings.value('last_selected_account', first_account_name)

    def save_config(self):
        if self.element_widgets:
            element_config_data = self.get_table_data()
        else:
            element_config_data = self.element_config

        data = {
            "LOGIN_URL": self.url_input.text(),
            "ORG_CODE": self.org_code_input.text(),
            "SKU_FILE_PATH": self.file_path_input.text(),
            "ACCOUNTS": self.all_accounts,
            "ELEMENT_CONFIG": element_config_data
        }

        if config_manager.save_config(data):
            self.log("配置已保存。", "green")
        else:
            QMessageBox.critical(self, "保存错误", "保存配置文件时出错。请查看日志。")

        if self.runtime_selected_account_name:
            self.config_settings.setValue('last_selected_account', self.runtime_selected_account_name)

        self.config_settings.setValue('url', self.url_input.text())
        self.config_settings.setValue('org_code', self.org_code_input.text())
        self.config_settings.setValue('sku_file_path', self.file_path_input.text())
        self.config_settings.setValue('start_point', self.start_point_combo.currentText())
        self.config_settings.setValue('headless', str(self.headless_checkbox.isChecked()).lower())
        self.config_settings.setValue('run_mode', self.mode_combo.currentText())

    # ... (UI 辅助方法省略) ...
    def load_account_profiles_to_ui(self):
        for i in reversed(range(self.account_list_layout.count())):
            widget_to_remove = self.account_list_layout.itemAt(i).widget()
            if widget_to_remove: widget_to_remove.setParent(None)
        if self.all_accounts:
            for account in self.all_accounts:
                btn = QPushButton(account["name"])
                btn.setProperty("account_name", account["name"])
                btn.clicked.connect(lambda checked, name=account["name"]: self.select_account_profile_by_name(name))
                self.account_list_layout.addWidget(btn)
            self.select_account_profile_by_name(self.runtime_selected_account_name, initial_load=True)
            self.delete_account_button.setEnabled(True)
        else:
            no_account_label = QLabel("无可用账号档案。")
            self.account_list_layout.addWidget(no_account_label)
            self.account_name_input.clear()
            self.username_input.clear()
            self.password_input.clear()
            self.delete_account_button.setEnabled(False)
            self.runtime_selected_account_name = ''

    def select_account_profile_by_name(self, name, initial_load=False):
        selected_account = None
        for account in self.all_accounts:
            if account["name"] == name: selected_account = account; break
        for i in range(self.account_list_layout.count()):
            widget = self.account_list_layout.itemAt(i).widget()
            if isinstance(widget, QPushButton):
                if widget.property("account_name") == name:
                    widget.setStyleSheet("background-color: #AECBFA; font-weight: bold;")
                else:
                    widget.setStyleSheet("")
        if selected_account:
            self.account_name_input.setText(selected_account.get("name", ""))
            self.username_input.setText(selected_account.get("username", ""))
            self.password_input.setText(selected_account.get("password", ""))
            self.delete_account_button.setEnabled(True)
            self.runtime_selected_account_name = name
            if not initial_load: self.save_config()
        else:
            self.account_name_input.clear()
            self.username_input.clear()
            self.password_input.clear()
            self.delete_account_button.setEnabled(False)
            self.runtime_selected_account_name = ''

    def save_current_account(self):
        name = self.account_name_input.text().strip()
        username = self.username_input.text().strip()
        password = self.password_input.text().strip()
        if not name or not username or not password:
            QMessageBox.warning(self, "保存失败", "所有字段都不能为空。")
            return
        existing_index = -1
        for i, account in enumerate(self.all_accounts):
            if account["name"] == name: existing_index = i; break
        new_account = {"name": name, "username": username, "password": password}
        if existing_index != -1:
            self.all_accounts[existing_index] = new_account
            self.log(f"账号档案 '{name}' 已更新。", "blue")
        else:
            self.all_accounts.append(new_account)
            self.log(f"账号档案 '{name}' 已新增。", "blue")
        self.save_config()
        self.load_account_profiles_to_ui()
        self.select_account_profile_by_name(name)

    def delete_current_account(self):
        current_name = self.account_name_input.text().strip()
        if not current_name: return
        if QMessageBox.question(self, '确认删除', f"确定要删除 '{current_name}' 吗?",
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            self.all_accounts = [acc for acc in self.all_accounts if acc["name"] != current_name]
            self.log(f"已删除。", "blue")
            self.runtime_selected_account_name = ''
            self.save_config()
            self.load_account_profiles_to_ui()
            if self.all_accounts:
                self.select_account_profile_by_name(self.all_accounts[0]['name'])
            else:
                self.select_account_profile_by_name('')

    def open_file_dialog(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "选择文件", "", "Excel/CSV (*.xlsx *.csv);;All (*)")
        if file_path:
            self.file_path_input.setText(file_path)
            self.save_config()

    def create_operation_page(self):
        op_page = QWidget()
        self.tab_widget.addTab(op_page, "操作执行")
        layout = QFormLayout(op_page)
        layout.addRow(QLabel("--- 账号档案配置 ---"))
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        account_list_widget = QWidget()
        self.account_list_layout = QVBoxLayout(account_list_widget)
        self.account_list_layout.setAlignment(Qt.AlignTop)
        scroll_area.setWidget(account_list_widget)
        scroll_area.setMaximumHeight(100)
        layout.addRow(scroll_area)
        self.account_name_input = QLineEdit()
        layout.addRow("档案名称:", self.account_name_input)
        self.username_input = QLineEdit()
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)
        layout.addRow("账号:", self.username_input)
        layout.addRow("密码:", self.password_input)
        account_btns = QHBoxLayout()
        self.save_account_button = QPushButton("保存档案")
        self.save_account_button.clicked.connect(self.save_current_account)
        self.delete_account_button = QPushButton("删除档案")
        self.delete_account_button.clicked.connect(self.delete_current_account)
        account_btns.addWidget(self.save_account_button)
        account_btns.addWidget(self.delete_account_button)
        layout.addRow(account_btns)
        layout.addRow(QLabel("--- 运行参数 ---"))
        self.url_input = QLineEdit(self.runtime_url)
        layout.addRow("目标 URL:", self.url_input)
        self.org_code_input = QLineEdit(self.runtime_org_code)
        layout.addRow("组织代码:", self.org_code_input)
        path_layout = QHBoxLayout()
        self.file_path_input = QLineEdit(self.sku_file_path)
        self.file_path_input.setReadOnly(True)
        f_btn = QPushButton("...")
        f_btn.clicked.connect(self.open_file_dialog)
        f_btn.setMaximumWidth(30)
        path_layout.addWidget(self.file_path_input)
        path_layout.addWidget(f_btn)
        layout.addRow("SKU 文件:", path_layout)
        self.start_point_combo = QComboBox()
        self.start_point_combo.addItems(["完整流程 (从登录开始)", "跳过登录/组织选择 (从导航开始)"])
        idx = self.start_point_combo.findText(self.runtime_start_point)
        if idx != -1: self.start_point_combo.setCurrentIndex(idx)
        layout.addRow("起始点:", self.start_point_combo)
        self.mode_combo = QComboBox()
        self.mode_combo.addItems(["模式一：自动导出 (按分类)", "模式二：人工筛选 (抓取所有)"])
        mode_idx = self.mode_combo.findText(self.runtime_run_mode)
        if mode_idx != -1: self.mode_combo.setCurrentIndex(mode_idx)
        layout.addRow("运行模式:", self.mode_combo)
        self.headless_checkbox = QCheckBox("后台静默运行 (无浏览器窗口)")
        self.headless_checkbox.setChecked(self.runtime_headless)
        layout.addRow("", self.headless_checkbox)
        self.execute_button = QPushButton("开始执行")
        self.execute_button.clicked.connect(self.start_automation)
        self.save_button = QPushButton("保存配置")
        self.save_button.clicked.connect(self.save_config)
        btns = QHBoxLayout()
        btns.addWidget(self.save_button)
        btns.addWidget(self.execute_button)
        layout.addRow(btns)
        self.log_output = QTextEdit()
        self.log_output.setReadOnly(True)
        layout.addRow(QLabel("日志:"), self.log_output)
        self.load_account_profiles_to_ui()

    def log(self, message, color="black"):
        fmt = QTextCharFormat()
        if color == "red":
            fmt.setForeground(Qt.red)
        elif color == "green":
            fmt.setForeground(Qt.green)
        elif color == "blue":
            fmt.setForeground(Qt.blue)
        else:
            fmt.setForeground(Qt.black)
        cursor = self.log_output.textCursor()
        cursor.movePosition(QTextCursor.End)
        cursor.insertText(f"{message}\n", fmt)
        self.log_output.ensureCursorVisible()
        QApplication.processEvents()

    def _read_skus_from_file(self, file_path):
        if not os.path.exists(file_path):
            self.log(f"ERROR: 文件未找到: {file_path}", "red")
            return []
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            skus = []
            for i, row in enumerate(ws.iter_rows(min_col=1, max_col=1, values_only=True)):
                if i == 0: continue
                if row[0]: skus.append(str(row[0]).strip())
            return skus
        except Exception as e:
            self.log(f"读取错误: {e}", "red")
            return []

    def _smart_find_element(self, driver, wait, element_name, parsed_config):
        if element_name not in parsed_config:
            raise KeyError(f"未找到元素配置: {element_name}")
        config = parsed_config[element_name]
        locator = config['locator_tuple']
        position = config.get('position', '当前元素')
        try:
            index = int(config.get('index', 1))
        except:
            index = 1
        target_element = None
        if index > 1:
            def elements_count_enough(d):
                eles = d.find_elements(*locator)
                return eles if len(eles) >= index else False

            found_elements = wait.until(elements_count_enough)
            base_element = found_elements[index - 1]
        else:
            base_element = wait.until(EC.presence_of_element_located(locator))
        if position == "当前元素":
            target_element = base_element
        elif position == "父元素":
            target_element = base_element.find_element(By.XPATH, "./..")
        elif position == "子元素":
            target_element = base_element.find_element(By.XPATH, "./*[1]")
        elif position == "上一个":
            target_element = base_element.find_element(By.XPATH, "preceding-sibling::*[1]")
        elif position == "下一个":
            target_element = base_element.find_element(By.XPATH, "following-sibling::*[1]")
        else:
            target_element = base_element
        return target_element

    def _update_request_session(self, driver):
        self.log("正在同步浏览器 Cookie...", "blue")
        selenium_cookies = driver.get_cookies()
        self.image_session.cookies.clear()
        self.image_session.headers.update({
            "User-Agent": driver.execute_script("return navigator.userAgent")
        })
        for cookie in selenium_cookies:
            self.image_session.cookies.set(cookie['name'], cookie['value'])
        self.log("Cookie 同步完成。", "green")

    def _execute_login_flow(self, driver, wait, parsed_config, url, username, password, org_code):
        self.log(f"1. 访问 URL: {url}...")
        driver.get(url)
        self.log(f"2. 输入账号和密码...")
        ele_user = self._smart_find_element(driver, wait, '账号输入框', parsed_config)
        ele_user.send_keys(username)
        ele_pass = self._smart_find_element(driver, wait, '密码输入框', parsed_config)
        ele_pass.send_keys(password)
        self.log("3. 点击登录按钮...")
        ele_login = self._smart_find_element(driver, wait, '登录按钮', parsed_config)
        ele_login.click()
        self.log("4. 等待组织选择弹窗...")
        self._smart_find_element(driver, wait, '组织选择弹窗', parsed_config)
        self.log("组织选择弹窗已出现。", "green")
        self.log(f"5. 输入组织代码: {org_code}")
        ele_org_input = self._smart_find_element(driver, wait, '组织输入框', parsed_config)
        ele_org_input.send_keys(org_code)
        org_item_config = parsed_config['组织列表项'].copy()
        original_locator = org_item_config['locator_tuple']
        if '156' in original_locator[1]:
            new_value = original_locator[1].replace('156', org_code)
            org_item_config['locator_tuple'] = (original_locator[0], new_value)
            parsed_config['组织列表项_动态'] = org_item_config
            self.log(f"6. 点击组织列表项...")
            ele_org_item = self._smart_find_element(driver, wait, '组织列表项_动态', parsed_config)
            driver.execute_script("arguments[0].click();", ele_org_item)
        else:
            ele_org_item = self._smart_find_element(driver, wait, '组织列表项', parsed_config)
            driver.execute_script("arguments[0].click();", ele_org_item)
        self.log("7. 点击确认登录按钮...")
        ele_confirm = self._smart_find_element(driver, wait, '确认登录按钮', parsed_config)
        ele_confirm.click()
        self.log("8. 等待跳转到首页...")
        wait.until(EC.url_contains("home_page"))
        self.log("完整登录和组织选择流程成功完成！", "green")

    def _execute_navigation_to_product_list(self, driver, wait, parsed_config):
        self.log("9. 执行导航流程: 鼠标悬停...")
        nav_icon = self._smart_find_element(driver, wait, '导航_商品主图标', parsed_config)
        wait.until(EC.visibility_of(nav_icon))
        ActionChains(driver).move_to_element(nav_icon).perform()
        self.log("悬停成功，等待子菜单...")
        time.sleep(0.5)
        self.log("10. 点击 '分销商品列表'...")
        nav_link = self._smart_find_element(driver, wait, '导航_分销商品列表', parsed_config)
        try:
            short_wait = WebDriverWait(driver, 5)
            short_wait.until(EC.visibility_of(nav_link))
            short_wait.until(EC.element_to_be_clickable(nav_link)).click()
        except TimeoutException:
            self.log("子菜单未显示，尝试 JS 强制点击...", "blue")
            driver.execute_script("arguments[0].click();", nav_link)
        self.log("11. 移开鼠标...", "blue")
        try:
            body = driver.find_element(By.TAG_NAME, "body")
            ActionChains(driver).move_to_element_with_offset(body, 0, 0).perform()
        except:
            pass
        self.log("12. 导航完成。", "green")

    def _execute_product_search(self, driver, wait, parsed_config, sku_value):
        self.log("12. 执行商品查询流程...")
        self.log(f"13. 输入 SKU: {sku_value}")
        inp = self._smart_find_element(driver, wait, 'product_list_sku_input', parsed_config)
        inp.clear()
        inp.send_keys(sku_value)
        self.log("14. 点击 '查询' 按钮...")
        btn = self._smart_find_element(driver, wait, 'product_list_search_button', parsed_config)
        try:
            btn.click()
        except:
            self.log("普通点击失败，转为 JS 点击...", "blue")
            driver.execute_script("arguments[0].click();", btn)
        self.log("查询指令已发送，等待结果刷新...", "blue")
        time.sleep(2)

        try:
            sku_locator = (By.XPATH, f"//*[contains(text(), '{sku_value}')]")
            wait.until(EC.presence_of_element_located(sku_locator))
            self.log(f"校验通过：页面已显示 SKU {sku_value}", "green")
        except TimeoutException:
            self.log(f"⚠️ 警告：等待 5 秒后页面仍未显示 SKU {sku_value}，可能查询无结果。", "red")
            time.sleep(2)

    def _capture_detail_info(self, driver, wait, parsed_config, mode="auto"):
        self.log("15. 抓取数据流程开始...")
        btn_detail = self._smart_find_element(driver, wait, 'product_list_view_detail_button', parsed_config)
        driver.execute_script("arguments[0].click();", btn_detail)
        self.log("17. 等待详情弹窗加载...")
        popup = self._smart_find_element(driver, wait, 'detail_popup_dialog', parsed_config)
        wait.until(EC.visibility_of(popup))
        self.log("等待图片加载...", "blue")
        time.sleep(1)
        try:
            driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", popup)
        except:
            pass
        time.sleep(1)
        self.log("18. 正在解析数据...", "blue")

        raw_info = {}
        ordered_keys = []
        try:
            table_element = popup.find_element(By.TAG_NAME, "table")
            rows = table_element.find_elements(By.TAG_NAME, "tr")
            for row in rows:
                cells = row.find_elements(By.TAG_NAME, "td")
                if len(cells) >= 2:
                    key = cells[0].text.strip().replace(':', '').replace('：', '')
                    if "FAQ" in key.upper(): continue
                    value = cells[1].text.strip()
                    raw_info[key] = value
                    ordered_keys.append(key)
        except Exception as e:
            self.log(f"表格解析出错: {e}", "black")

        final_text_info = self._process_text_info(raw_info, ordered_keys)
        image_data = {}
        all_images = []

        if "按分类" in mode:
            image_categories = ['全家福', '主图', '细节图', '尺寸图', '卖点图', '其他图']
            for category in image_categories:
                try:
                    xpath = f".//h5[contains(., '{category}')]/following-sibling::div[1]//img"
                    imgs = popup.find_elements(By.XPATH, xpath)
                    if imgs:
                        src = imgs[0].get_attribute('src')
                        if src and "empty" not in src:
                            image_data[category] = src
                        else:
                            image_data[category] = ""
                    else:
                        image_data[category] = ""
                except:
                    image_data[category] = ""
        else:
            try:
                imgs = popup.find_elements(By.TAG_NAME, "img")
                for img in imgs:
                    src = img.get_attribute('src')
                    if src and len(src) > 20 and "empty" not in src:
                        all_images.append(src)
            except Exception as e:
                self.log(f"图片抓取出错: {e}", "red")

        close_btn = self._smart_find_element(driver, wait, 'detail_close_button', parsed_config)
        try:
            close_btn.click()
        except:
            driver.execute_script("arguments[0].click();", close_btn)
        try:
            wait.until(EC.invisibility_of_element(popup))
        except:
            pass
        self.log("详情信息抓取完成。", "green")
        return {
            "text_info": final_text_info,
            "image_data": image_data,
            "all_images": all_images
        }

    def _process_text_info(self, raw_info, ordered_keys):
        new_info = {}
        desc_parts = []
        merge_keys = ['Specification', 'Package List', 'Note']
        for key in ordered_keys:
            if key in merge_keys:
                val = raw_info.get(key)
                if val: desc_parts.append(f"{key}: {val}")
                continue
            if key == 'Features':
                features_text = raw_info.get(key, "")
                parts = re.split(r'\d+、', features_text)
                clean_parts = [p.strip() for p in parts if p.strip()]
                for i in range(5):
                    key_name = f"五点{i + 1}"
                    val = clean_parts[i] if i < len(clean_parts) else ""
                    new_info[key_name] = val
                continue
            new_info[key] = raw_info.get(key, "")
        if desc_parts:
            new_info['描述'] = "\n".join(desc_parts)
        return new_info

    def start_automation(self):
        if not SELENIUM_AVAILABLE:
            self.log("ERROR: Selenium 未安装。", "red")
            return
        self.log("--- 自动化执行开始 ---", "blue")
        self.save_config()
        url = self.url_input.text()
        username = self.username_input.text()
        password = self.password_input.text()
        org_code = self.org_code_input.text()
        sku_file_path = self.file_path_input.text()
        start_point = self.start_point_combo.currentText()
        run_mode = self.mode_combo.currentText()
        is_headless = self.headless_checkbox.isChecked()

        if start_point == "完整流程 (从登录开始)" and (not username or not password):
            self.log(f"FATAL ERROR: 账号密码不能为空。", "red")
            return
        config_data = self.get_table_data()
        parsed_config = {}
        for module_item in config_data:
            for element in module_item.get("elements", []):
                by, value = LocatorParser.parse(element["locator"])
                if by and value:
                    parsed_config[element["name"]] = {
                        "locator_tuple": (by, value),
                        "position": element.get("position", "当前元素"),
                        "index": element.get("index", "1")
                    }
        sku_list = self._read_skus_from_file(sku_file_path)
        if not sku_list:
            self.log("FATAL ERROR: 读取 SKU 失败。", "red")
            return
        self.log(f"成功读取 {len(sku_list)} 个 SKU。", "blue")
        try:
            service = Service()
            options = EdgeOptions()
            if is_headless:
                self.log("启用静默模式 (Headless)...", "blue")
                options.add_argument("--headless")
                options.add_argument("--disable-gpu")
                options.add_argument("--window-size=1920,1080")
            driver = webdriver.Edge(service=service, options=options)
            if not is_headless: driver.maximize_window()
            self.log("Edge 浏览器启动成功。", "green")
        except Exception as e:
            self.log(f"FATAL ERROR: 浏览器启动失败: {e}", "red")
            return

        all_results = []
        try:
            wait = WebDriverWait(driver, 15)
            if start_point == "完整流程 (从登录开始)":
                self._execute_login_flow(driver, wait, parsed_config, url, username, password, org_code)
            elif start_point == "跳过登录/组织选择 (从导航开始)":
                self.log("直接从导航开始...", "blue")
                driver.get(url.split('#')[0] + '#/product/distribution_list')

            self._update_request_session(driver)
            self._execute_navigation_to_product_list(driver, wait, parsed_config)

            for sku_idx, current_sku in enumerate(sku_list):
                self.log(f"\n--- 开始处理 SKU: {current_sku} ({sku_idx + 1}/{len(sku_list)}) ---", "blue")
                try:
                    self._execute_product_search(driver, wait, parsed_config, current_sku)
                    capture_result = self._capture_detail_info(driver, wait, parsed_config, mode=run_mode)
                    result_row = {"SKU": current_sku}
                    result_row.update(capture_result)
                    all_results.append(result_row)
                    self.log(f"SKU {current_sku} 处理成功。", "green")
                    time.sleep(1)
                except Exception as e:
                    error_msg = str(e).split('\n')[0]
                    self.log(f"SKU {current_sku} 处理出错 (已跳过): {error_msg}", "red")
                    all_results.append({"SKU": current_sku, "ERROR": "处理失败", "Details": error_msg})
                    try:
                        driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE)
                    except:
                        pass

            self.log("\n自动化测试运行结束。", "green")
            if "模式一" in run_mode:
                self._handle_mode_one_export(all_results, sku_file_path)
            else:
                self._handle_mode_two_dialog(all_results, sku_file_path)

        except Exception as e:
            self.log(f"执行中发生严重错误 (流程终止): {e}", "red")
            import traceback
            self.log(traceback.format_exc(), "black")
        finally:
            self.log("结束。", "blue")

    def _handle_mode_one_export(self, all_results, source_path):
        self.log("正在执行自动导出...", "blue")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_dir = os.path.dirname(source_path) if source_path else os.getcwd()
        filename = f"自动抓取结果_{timestamp}.xlsx"
        save_path = os.path.join(base_dir, filename)
        workbook = xlsxwriter.Workbook(save_path)
        worksheet = workbook.add_worksheet()
        text_keys = []
        for res in all_results:
            for k in res.get('text_info', {}).keys():
                if k not in text_keys: text_keys.append(k)
        img_keys = ['全家福', '主图', '细节图', '尺寸图', '卖点图', '其他图']
        headers = ["SKU"] + text_keys + img_keys + ["ERROR"]
        header_format = workbook.add_format({'bold': True, 'bg_color': '#D7E4BC', 'border': 1})
        center_format = workbook.add_format({'valign': 'vcenter', 'border': 1})
        wrap_format = workbook.add_format({'valign': 'vcenter', 'border': 1, 'text_wrap': True})
        for idx, h in enumerate(headers):
            worksheet.write(0, idx, h, header_format)
            if h in img_keys:
                worksheet.set_column(idx, idx, 40)
            elif "描述" in h:
                worksheet.set_column(idx, idx, 40)
            else:
                worksheet.set_column(idx, idx, 20)
        current_row = 1
        for res in all_results:
            worksheet.write(current_row, 0, res['SKU'], center_format)
            col_offset = 1
            for k in text_keys:
                fmt = wrap_format if k == "描述" else center_format
                worksheet.write(current_row, col_offset, res.get('text_info', {}).get(k, ""), fmt)
                col_offset += 1
            img_start_col = col_offset
            for idx, k in enumerate(img_keys):
                worksheet.write(current_row, img_start_col + idx, res.get('image_data', {}).get(k, ""), center_format)
            err_col = img_start_col + len(img_keys)
            if 'ERROR' in res:
                worksheet.write(current_row, err_col, f"{res['ERROR']} - {res.get('Details', '')}", center_format)
            else:
                worksheet.write(current_row, err_col, "", center_format)
            worksheet.set_row(current_row, 20)
            current_row += 1
        try:
            workbook.close()
            self.log(f"导出成功! 文件路径: {save_path}", "green")
            QMessageBox.information(self, "完成", f"自动抓取完成。\n文件已保存至: {save_path}")
        except Exception as e:
            self.log(f"导出失败: {e}", "red")

    def _handle_mode_two_dialog(self, all_results, source_path):
        self.log("打开筛选窗口...", "blue")
        dialog = ImageSelectorDialog(all_results, source_path, self.image_session, self)
        dialog.exec_()

    def create_config_page(self):
        config_page = QWidget()
        self.tab_widget.addTab(config_page, "元素配置")
        layout = QVBoxLayout(config_page)
        help_text = ("<b>配置指南:</b><br>"
                     "1. <b>定位字符串:</b> 支持 XPath, class, 属性, 纯文本。<br>"
                     "2. <b>位置:</b> 选择相对于定位元素的导航方向 (如父/子元素)。<br>"
                     "3. <b>Index:</b> 如果定位到多个元素，选择第几个 (默认1)。")
        layout.addWidget(QLabel(help_text))
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        config_container = QWidget()
        container_layout = QVBoxLayout(config_container)
        self.element_widgets = {}
        for module_item in self.element_config:
            module_name = module_item["module"]
            group_box = QGroupBox(module_name)
            group_layout = QFormLayout(group_box)
            for element in module_item["elements"]:
                name = element["name"]
                row_widget = QWidget()
                row_layout = QHBoxLayout(row_widget)
                row_layout.setContentsMargins(0, 0, 0, 0)
                locator_input = QLineEdit(element["locator"])
                locator_input.setPlaceholderText("定位字符串")
                position_combo = QComboBox()
                position_combo.addItems(["当前元素", "父元素", "子元素", "上一个", "下一个"])
                position_combo.setCurrentText(element.get("position", "当前元素"))
                position_combo.setFixedWidth(80)
                index_input = QLineEdit(str(element.get("index", "1")))
                index_input.setPlaceholderText("Idx")
                from PyQt5.QtGui import QIntValidator
                index_input.setValidator(QIntValidator(1, 999))
                index_input.setFixedWidth(30)
                row_layout.addWidget(locator_input, stretch=3)
                row_layout.addWidget(position_combo, stretch=1)
                row_layout.addWidget(index_input, stretch=0)
                self.element_widgets[name] = {
                    "locator": locator_input,
                    "position": position_combo,
                    "index": index_input
                }
                group_layout.addRow(name, row_widget)
            container_layout.addWidget(group_box)
        scroll_area.setWidget(config_container)
        layout.addWidget(scroll_area)
        save_config_button = QPushButton("保存元素配置")
        save_config_button.clicked.connect(self.save_config)
        layout.addWidget(save_config_button)

    def get_table_data(self):
        updated_config = []
        for module_item in self.element_config:
            new_module = {"module": module_item["module"], "elements": []}
            for element in module_item["elements"]:
                name = element["name"]
                widgets = self.element_widgets.get(name)
                if widgets:
                    new_element = {
                        "name": name,
                        "locator": widgets["locator"].text(),
                        "position": widgets["position"].currentText(),
                        "index": widgets["index"].text() or "1"
                    }
                    new_module["elements"].append(new_element)
                else:
                    new_module["elements"].append(element)
            updated_config.append(new_module)
        return updated_config


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = AutomationToolUI()
    window.show()
    sys.exit(app.exec_())