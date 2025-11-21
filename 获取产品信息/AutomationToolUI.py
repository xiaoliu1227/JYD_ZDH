import re
import sys
import os
import time
import openpyxl
import requests
import gc
from datetime import datetime
from io import BytesIO
from PIL import Image as PilImage

from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QTabWidget, QFormLayout, QScrollArea, QComboBox,
                             QLineEdit, QPushButton, QLabel, QDialog, QGroupBox,
                             QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox, QTextEdit, QFileDialog,
                             QCheckBox, QGridLayout)
from PyQt5.QtCore import Qt, QSettings, QThread, pyqtSignal, QWaitCondition, QMutex, QRect
from PyQt5.QtGui import QFont, QTextCharFormat, QTextCursor, QPixmap, QPainter, QPen, QBrush

from config_manager import config_manager
from edge_automation_tool import LocatorParser
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

try:
    from selenium import webdriver
    from selenium.webdriver.edge.service import Service
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.action_chains import ActionChains
    from selenium.common.exceptions import TimeoutException, NoSuchElementException
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.common.by import By
    from selenium.webdriver.edge.options import Options as EdgeOptions

    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False

# 固定表头
FIXED_TEXT_HEADERS = [
    "商品目录", "尺寸(cm)", "重量(g)", "中文名", "是否带Logo",
    "商品属性", "采购交期", "Item Title", "Key Words",
    "五点1", "五点2", "五点3", "五点4", "五点5", "描述"
]

IMG_HEADERS_AUTO = [
    "全家福", "主图", "细节图", "尺寸图", "卖点图", "其他图",
    "备用图_1", "备用图_2", "备用图_3"
]
IMG_HEADERS_MANUAL = [f"图片链接_{i+1}" for i in range(9)]

# --- 1. 后台工作线程 ---
class AutomationWorker(QThread):
    log_signal = pyqtSignal(str, str)
    progress_signal = pyqtSignal(str)
    batch_ready_signal = pyqtSignal(list, object)
    finished_signal = pyqtSignal()
    error_signal = pyqtSignal(str)

    def __init__(self, config_data, sku_file_path, run_mode, start_point, is_headless, output_path):
        super().__init__()
        self.config_data = config_data
        self.sku_file_path = sku_file_path
        self.run_mode = run_mode
        self.start_point = start_point
        self.is_headless = is_headless
        self.output_path = output_path

        self.is_running = True
        self.is_paused = False
        self.mutex = QMutex()
        self.wait_condition = QWaitCondition()

        self.image_session = requests.Session()

    def stop(self):
        self.is_running = False
        self.resume()

    def resume(self):
        self.mutex.lock()
        self.is_paused = False
        self.wait_condition.wakeAll()
        self.mutex.unlock()

    def wait_for_user(self):
        self.mutex.lock()
        self.is_paused = True
        self.wait_condition.wait(self.mutex)
        self.mutex.unlock()

    def run(self):
        driver = None
        try:
            sku_list = self._read_skus_from_file(self.sku_file_path)
            if not sku_list:
                self.error_signal.emit("SKU 文件为空或无法读取")
                return

            processed_skus = self._get_processed_skus(self.output_path)
            self.log_signal.emit(f"检测到 {len(processed_skus)} 个已完成 SKU，将自动跳过。", "blue")

            driver = self._init_driver()
            wait = WebDriverWait(driver, 15)

            # 【修复点 1】正确调用解析方法
            parsed_config = self._parse_config()

            # 执行登录/导航
            if "完整流程" in self.start_point:
                self._execute_login(driver, wait, parsed_config)
            else:
                # 从配置中获取 URL
                login_url = self.config_data.get('LOGIN_URL', 'https://saaserp-pos.yibainetwork.com/#/login_page')
                target_url = login_url.split('#')[0] + '#/product/distribution_list'
                driver.get(target_url)
                self.log_signal.emit("已跳转至列表页 (跳过登录)。", "blue")

            self._update_session_cookies(driver)
            self._execute_navigation(driver, wait, parsed_config)

            batch_size = 50 if "模式一" in self.run_mode else 20
            current_batch = []

            for i, sku in enumerate(sku_list):
                if not self.is_running: break

                if sku in processed_skus:
                    continue

                self.log_signal.emit(f"正在处理 ({i + 1}/{len(sku_list)}): {sku}", "black")

                try:
                    self._execute_search(driver, wait, parsed_config, sku)
                    data = self._capture_data(driver, wait, parsed_config, self.run_mode)

                    row_data = {"SKU": sku}
                    row_data.update(data)
                    current_batch.append(row_data)

                    self.log_signal.emit(f"SKU {sku} 获取成功，暂存入批次。", "green")

                    if len(current_batch) >= batch_size:
                        self.log_signal.emit(f"批次已满 ({batch_size})，等待主程序处理...", "blue")
                        self.batch_ready_signal.emit(current_batch, self.image_session)
                        self.wait_for_user()

                        current_batch = []
                        gc.collect()

                        if not self.is_running: break

                    time.sleep(1)

                except Exception as e:
                    err_msg = str(e).split('\n')[0]
                    self.log_signal.emit(f"SKU {sku} 出错: {err_msg}", "red")
                    current_batch.append({"SKU": sku, "ERROR": "处理失败", "Details": err_msg})
                    self._recover_browser(driver)

            if current_batch and self.is_running:
                self.log_signal.emit(f"处理剩余 {len(current_batch)} 条数据...", "blue")
                self.batch_ready_signal.emit(current_batch, self.image_session)
                self.wait_for_user()

            self.finished_signal.emit()

        except Exception as e:
            import traceback
            self.error_signal.emit(f"线程异常: {e}\n{traceback.format_exc()}")
        finally:
            if driver:
                driver.quit()

    # --- 辅助方法 ---
    def _read_skus_from_file(self, path):
        if not os.path.exists(path): return []
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            skus = []
            for i, row in enumerate(ws.iter_rows(min_col=1, max_col=1, values_only=True)):
                if i == 0: continue
                if row[0]: skus.append(str(row[0]).strip())
            return skus
        except:
            return []

    def _get_processed_skus(self, output_path):
        if not os.path.exists(output_path): return set()
        try:
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            skus = set()
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
                if row[0]: skus.add(str(row[0]).strip())
            return skus
        except:
            return set()

    def _init_driver(self):
        service = Service()
        options = EdgeOptions()
        if self.is_headless:
            options.add_argument("--headless")
            options.add_argument("--disable-gpu")
            options.add_argument("--window-size=1920,1080")
        driver = webdriver.Edge(service=service, options=options)
        if not self.is_headless: driver.maximize_window()
        return driver

    def _parse_config(self):
        """解析配置，从字典中提取元素列表"""
        parsed = {}
        # 【核心修复】明确获取 'ELEMENT_CONFIG' 列表，而不是遍历 config_data 字典
        element_modules = self.config_data.get('ELEMENT_CONFIG', [])

        for module in element_modules:
            for ele in module['elements']:
                by, val = LocatorParser.parse(ele['locator'])
                if by:
                    parsed[ele['name']] = {
                        'locator': (by, val),
                        'position': ele.get('position', '当前元素'),
                        'index': int(ele.get('index', 1))
                    }
        return parsed

    def _find(self, driver, wait, name, config):
        cfg = config.get(name)
        if not cfg: raise KeyError(f"未配置: {name}")

        locator = cfg['locator']
        index = cfg['index']
        position = cfg['position']

        target = None
        if index > 1:
            def count_ok(d):
                eles = d.find_elements(*locator)
                return eles if len(eles) >= index else False

            found = wait.until(count_ok)
            base = found[index - 1]
        else:
            base = wait.until(EC.presence_of_element_located(locator))

        if position == "父元素":
            target = base.find_element(By.XPATH, "./..")
        elif position == "子元素":
            target = base.find_element(By.XPATH, "./*[1]")
        elif position == "上一个":
            target = base.find_element(By.XPATH, "preceding-sibling::*[1]")
        elif position == "下一个":
            target = base.find_element(By.XPATH, "following-sibling::*[1]")
        else:
            target = base
        return target

    def _update_session_cookies(self, driver):
        cookies = driver.get_cookies()
        self.image_session.cookies.clear()
        self.image_session.headers.update({"User-Agent": driver.execute_script("return navigator.userAgent")})
        for c in cookies:
            self.image_session.cookies.set(c['name'], c['value'])

    def _execute_login(self, driver, wait, config):
        self.log_signal.emit("开始登录...", "black")
        url = self.config_data.get('LOGIN_URL', "")

        # 下面复用之前的登录逻辑步骤，只是改用 self._find
        driver.get(url)

        # 填账号
        ele = self._find(driver, wait, '账号输入框', config)
        # 这里我们没法直接拿到 username，需要从 config_data 或者其他地方传。
        # 为了简化，我们在 AutomationToolUI 中构建 worker_config 时把 username 也放进去
        # 或者直接操作元素
        # 注意：由于账号密码是存在 settings 里的，这里需要调整一下传参结构
        # 临时方案：在 start_automation 构建 worker_config 时加入 'USERNAME', 'PASSWORD', 'ORG_CODE'

        ele.send_keys(self.config_data.get('USERNAME', ''))

        ele = self._find(driver, wait, '密码输入框', config)
        ele.send_keys(self.config_data.get('PASSWORD', ''))

        ele = self._find(driver, wait, '登录按钮', config)
        ele.click()

        self._find(driver, wait, '组织选择弹窗', config)

        ele = self._find(driver, wait, '组织输入框', config)
        org_code = self.config_data.get('ORG_CODE', '156')
        ele.send_keys(org_code)

        # 组织列表项动态处理
        try:
            # 尝试直接点击列表项（假设只有一个匹配）
            ele = self._find(driver, wait, '组织列表项', config)
            driver.execute_script("arguments[0].click();", ele)
        except:
            # 如果需要替换文本
            # 这里简单处理，假设配置里的定位器是通用的
            pass

        ele = self._find(driver, wait, '确认登录按钮', config)
        ele.click()

        wait.until(EC.url_contains("home_page"))
        self.log_signal.emit("登录成功。", "green")

    def _execute_navigation(self, driver, wait, config):
        self.log_signal.emit("导航中...", "black")
        nav_icon = self._find(driver, wait, '导航_商品主图标', config)
        wait.until(EC.visibility_of(nav_icon))
        ActionChains(driver).move_to_element(nav_icon).perform()
        time.sleep(0.5)

        nav_link = self._find(driver, wait, '导航_分销商品列表', config)
        try:
            short_wait = WebDriverWait(driver, 5)
            short_wait.until(EC.visibility_of(nav_link))
            short_wait.until(EC.element_to_be_clickable(nav_link)).click()
        except:
            self.log_signal.emit("JS 强制点击菜单...", "blue")
            driver.execute_script("arguments[0].click();", nav_link)

        try:
            ActionChains(driver).move_to_element_with_offset(driver.find_element(By.TAG_NAME, "body"), 0, 0).perform()
        except:
            pass

    def _execute_search(self, driver, wait, config, sku):
        self.log_signal.emit(f"查询 SKU: {sku}", "black")
        inp = self._find(driver, wait, 'product_list_sku_input', config)
        inp.clear()
        inp.send_keys(sku)
        btn = self._find(driver, wait, 'product_list_search_button', config)
        try:
            btn.click()
        except:
            driver.execute_script("arguments[0].click();", btn)

        time.sleep(2)
        try:
            wait.until(EC.presence_of_element_located((By.XPATH, f"//*[contains(text(), '{sku}')]")))
        except:
            self.log_signal.emit("等待结果超时，可能无数据", "red")
            time.sleep(2)

    def _capture_data(self, driver, wait, config, mode):
        # 1. 打开详情弹窗
        btn = self._find(driver, wait, 'product_list_view_detail_button', config)
        driver.execute_script("arguments[0].click();", btn)

        popup = self._find(driver, wait, 'detail_popup_dialog', config)
        wait.until(EC.visibility_of(popup))
        time.sleep(1)

        # 滚动到底部加载图片
        try:
            driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", popup)
        except:
            pass
        time.sleep(1)

        # --- 抓取文本信息 (保持不变) ---
        text_info = {}
        try:
            table = popup.find_element(By.TAG_NAME, "table")
            for row in table.find_elements(By.TAG_NAME, "tr"):
                cells = row.find_elements(By.TAG_NAME, "td")
                if len(cells) >= 2:
                    k = cells[0].text.strip().replace(':', '').replace('：', '')
                    if "FAQ" in k.upper(): continue
                    text_info[k] = cells[1].text.strip()
        except:
            pass

        # --- 【核心修改：图片抓取逻辑区分模式】 ---
        all_images = []  # 用于模式二（存所有图）
        categorized_images = {}  # 用于模式一（存分类图）

        if "模式一" in mode:
            # 定义需要的分类顺序
            categories = ["全家福", "主图", "细节图", "尺寸图", "卖点图", "其他图"]

            for cat in categories:
                img_url = ""  # 默认为空
                try:
                    # XPath 逻辑:
                    # 1. 找到包含特定文本(如"全家福")的 h5 标签
                    # 2. 找到该 h5 标签紧邻的下一个 div (/following-sibling::div[1])
                    # 3. 在该 div 下寻找 img 标签
                    xpath = f".//h5[contains(text(), '{cat}')]/following-sibling::div[1]//img"

                    # 在弹窗范围内查找
                    imgs = popup.find_elements(By.XPATH, xpath)

                    for img in imgs:
                        src = img.get_attribute("src")
                        # 简单的有效性检查
                        if src and len(src) > 20 and "empty" not in src:
                            img_url = src
                            break  # 找到第一张就停止，进入下一个分类
                except Exception as e:
                    # 某个分类找不到不影响其他分类
                    pass

                categorized_images[cat] = img_url
        else:
            # --- 模式二：保持原有的“抓取所有图片”逻辑 ---
            try:
                for img in popup.find_elements(By.TAG_NAME, "img"):
                    src = img.get_attribute("src")
                    if src and len(src) > 20 and "empty" not in src:
                        all_images.append(src)
            except:
                pass

        # 关闭弹窗
        close = self._find(driver, wait, 'detail_close_button', config)
        try:
            close.click()
        except:
            driver.execute_script("arguments[0].click();", close)
        try:
            wait.until(EC.invisibility_of_element(popup))
        except:
            pass

        # 返回结果增加 categorized_images 字段
        return {
            "text_info": text_info,
            "all_images": all_images,
            "categorized_images": categorized_images
        }

    def _recover_browser(self, driver):
        try:
            driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE)
        except:
            pass


# --- 2. 图片筛选弹窗 ---
class ImageSelectorDialog(QDialog):
    def __init__(self, batch_results, batch_index, image_session, parent=None):
        super().__init__(parent)
        self.batch_results = batch_results
        self.image_session = image_session
        self.processed_data = []
        self.action = "continue"

        self.setWindowTitle(f"筛选 - 第 {batch_index} 批")
        self.setGeometry(100, 100, 1200, 800)

        self.layout = QVBoxLayout(self)
        self.layout.addWidget(QLabel(f"本批次共 {len(batch_results)} 个 SKU。请筛选图片。"))

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        container = QWidget()
        self.container_layout = QVBoxLayout(container)

        self.sku_widgets = []
        for res in batch_results:
            if 'ERROR' in res: continue
            sku_box = SKUResultWidget(res, self.image_session)
            self.container_layout.addWidget(sku_box)
            self.sku_widgets.append(sku_box)

        container.setLayout(self.container_layout)
        scroll.setWidget(container)
        self.layout.addWidget(scroll)

        btn_layout = QHBoxLayout()

        btn_continue = QPushButton("保存并继续下一批")
        btn_continue.clicked.connect(self.on_continue)
        btn_continue.setStyleSheet("background-color: #4CAF50; color: white; font-size: 14px; height: 40px;")

        btn_stop = QPushButton("保存并停止运行")
        btn_stop.clicked.connect(self.on_stop)
        btn_stop.setStyleSheet("background-color: #f44336; color: white; font-size: 14px; height: 40px;")

        btn_layout.addWidget(btn_continue)
        btn_layout.addWidget(btn_stop)
        self.layout.addLayout(btn_layout)

    def _collect_data(self):
        self.processed_data = []
        for res in self.batch_results:
            if 'ERROR' in res:
                self.processed_data.append(res)

        for i, sku_widget in enumerate(self.sku_widgets):
            original_data = sku_widget.sku_data
            selected_images = sku_widget.get_selected_images()
            data = original_data.copy()
            data['images'] = selected_images
            self.processed_data.append(data)

    def on_continue(self):
        self._collect_data()
        self.action = "continue"
        self.accept()

    def on_stop(self):
        self._collect_data()
        self.action = "stop"
        self.accept()


# --- 3. 主窗口 ---
class AutomationToolUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Edge 自动化配置工具 (多线程稳健版)")
        self.setGeometry(100, 100, 950, 800)
        self.config_settings = QSettings('MyCompany', 'EdgeAutoTool')

        self.all_accounts = []
        self.element_config = []
        self.element_widgets = {}

        self.load_config()

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QVBoxLayout(self.central_widget)
        self.tab_widget = QTabWidget()
        self.main_layout.addWidget(self.tab_widget)
        self.create_operation_page()
        self.create_config_page()

        self.worker = None

    def _unify_element_config(self, code_structure, json_data):
        json_map = {}
        if json_data:
            for module_item in json_data:
                for element in module_item.get("elements", []):
                    json_map[element["name"]] = element
        unified_config = code_structure.copy()
        for module_item in unified_config:
            for element in module_item["elements"]:
                name = element["name"]
                if name in json_map:
                    element["locator"] = json_map[name].get("locator", "")
                    element["position"] = json_map[name].get("position", "当前元素")
                    element["index"] = json_map[name].get("index", "1")
                else:
                    if "position" not in element: element["position"] = "当前元素"
                    if "index" not in element: element["index"] = "1"
        return unified_config

    def load_config(self):
        config = config_manager.load_config()
        default_config = config_manager.default_config
        self.all_accounts = config.get("ACCOUNTS", [])
        self.element_config = self._unify_element_config(
            default_config.get("ELEMENT_CONFIG", []),
            config.get("ELEMENT_CONFIG_FROM_FILE", [])
        )
        self.runtime_url = self.config_settings.value('url', config.get("LOGIN_URL"))
        self.runtime_org_code = self.config_settings.value('org_code', config.get("ORG_CODE"))
        self.sku_file_path = self.config_settings.value('sku_file_path', config.get("SKU_FILE_PATH"))
        self.runtime_start_point = self.config_settings.value('start_point', '完整流程 (从登录开始)')
        self.runtime_headless = self.config_settings.value('headless', 'false') == 'true'
        self.runtime_run_mode = self.config_settings.value('run_mode', '模式一：自动导出 (按分类)')
        first_account_name = self.all_accounts[0]['name'] if self.all_accounts else ''
        self.runtime_selected_account_name = self.config_settings.value('last_selected_account', first_account_name)

    def save_config(self):
        if self.element_widgets:
            element_config_data = self.get_table_data()
        else:
            element_config_data = self.element_config
        data = {
            "LOGIN_URL": self.url_input.text(),
            "ORG_CODE": self.org_code_input.text(),
            "SKU_FILE_PATH": self.file_path_input.text(),
            "ACCOUNTS": self.all_accounts,
            "ELEMENT_CONFIG": element_config_data
        }
        config_manager.save_config(data)
        if self.runtime_selected_account_name:
            self.config_settings.setValue('last_selected_account', self.runtime_selected_account_name)
        self.config_settings.setValue('url', self.url_input.text())
        self.config_settings.setValue('org_code', self.org_code_input.text())
        self.config_settings.setValue('sku_file_path', self.file_path_input.text())
        self.config_settings.setValue('start_point', self.start_point_combo.currentText())
        self.config_settings.setValue('headless', str(self.headless_checkbox.isChecked()).lower())
        self.config_settings.setValue('run_mode', self.mode_combo.currentText())

    def start_automation(self):
        self.save_config()
        config_structure = self.get_table_data()

        # 【修复点 2】构建完整的配置字典，传入账号信息
        worker_config = {
            'LOGIN_URL': self.url_input.text(),
            'ORG_CODE': self.org_code_input.text(),
            'USERNAME': self.username_input.text(),
            'PASSWORD': self.password_input.text(),
            'ELEMENT_CONFIG': config_structure
        }

        self.worker = AutomationWorker(
            worker_config,
            self.file_path_input.text(),
            self.mode_combo.currentText(),
            self.start_point_combo.currentText(),
            self.headless_checkbox.isChecked(),
            output_path=""
        )

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_dir = os.path.dirname(self.file_path_input.text()) if self.file_path_input.text() else os.getcwd()
        output_path = os.path.join(base_dir, f"抓取结果_{timestamp}.xlsx")
        self.worker.output_path = output_path

        self.worker.log_signal.connect(self.log)
        self.worker.batch_ready_signal.connect(self.handle_batch_ready)
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.error_signal.connect(lambda msg: self.log(f"错误: {msg}", "red"))

        self.execute_button.setEnabled(False)
        self.stop_button.setEnabled(True)
        self.worker.start()

    def stop_automation(self):
        if self.worker and self.worker.isRunning():
            self.log("正在请求停止... 请等待当前 SKU 完成...", "red")
            self.worker.stop()
            self.stop_button.setEnabled(False)

    def on_finished(self):
        self.execute_button.setEnabled(True)
        self.stop_button.setEnabled(False)
        self.log("任务结束。", "blue")

    def handle_batch_ready(self, batch_data, session):
        final_rows = []

        if "模式一" in self.worker.run_mode:
            for item in batch_data:
                final_rows.append(self._format_row(item))
            self.log(f"自动保存批次 ({len(final_rows)} 条)...", "green")
        else:
            dialog = ImageSelectorDialog(batch_data, 0, session, self)
            result = dialog.exec_()

            if result == QDialog.Accepted:
                processed = dialog.processed_data
                for item in processed:
                    final_rows.append(self._format_row(item))

                if dialog.action == "stop":
                    self.worker.stop()
            else:
                self.worker.stop()

        self._append_to_excel(self.worker.output_path, final_rows)
        self.worker.resume()

    def _format_row(self, item):
        # 1. 获取原始抓取到的文本字典
        raw_info = item.get('text_info', {})

        # --- 文本处理逻辑 (保持不变) ---
        features_text = raw_info.get('Features', '')
        split_pattern = r'\d+[、]\s*'
        parts = re.split(split_pattern, features_text)
        feature_lines = [p.strip() for p in parts if p.strip()]

        five_points = [""] * 5
        for i in range(min(len(feature_lines), 5)):
            five_points[i] = feature_lines[i]

        desc_parts = []
        merge_fields = [('Specification', 'Specification'), ('Package List', 'Package List'), ('Note', 'Note')]
        for key, title in merge_fields:
            val = raw_info.get(key, '').strip()
            if val:
                desc_parts.append(f"{title}:\n{val}")
        full_description = "\n\n".join(desc_parts)

        data_mapping = {
            "五点1": five_points[0], "五点2": five_points[1], "五点3": five_points[2],
            "五点4": five_points[3], "五点5": five_points[4], "描述": full_description,
            "商品目录": raw_info.get('Category', raw_info.get('商品目录', '')),
            "中文名": raw_info.get('Chinese Name', raw_info.get('中文名', '')),
            "Item Title": raw_info.get('Item Title', raw_info.get('Title', '')),
            "尺寸(cm)": raw_info.get('Size', raw_info.get('尺寸(cm)', '')),
            "重量(g)": raw_info.get('Weight', raw_info.get('重量(g)', '')),
        }

        row = [item['SKU']]
        for h in FIXED_TEXT_HEADERS:
            row.append(data_mapping.get(h, raw_info.get(h, "")))

        # --- 【核心修改：图片列填充逻辑】 ---

        # 优先判断是否存在分类图片数据 (即模式一)
        if item.get('categorized_images'):
            # 按照固定顺序填入 Excel 的前 6 列图片位置
            categories = ["全家福", "主图", "细节图", "尺寸图", "卖点图", "其他图"]
            cat_imgs = item['categorized_images']

            for cat in categories:
                row.append(cat_imgs.get(cat, ""))  # 填入对应链接，没有则为空

            # 补齐剩余的空列 (假设总共要留9个图片位)
            for _ in range(9 - len(categories)):
                row.append("")

        else:
            # 模式二逻辑 (人工筛选后的列表，或者原始列表)
            images = []
            if 'images' in item:  # 人工筛选后的结果存放在 'images'
                images = item['images']
            elif 'all_images' in item:  # 原始抓取结果
                images = item['all_images']

            for i in range(9):
                if i < len(images):
                    row.append(images[i])
                else:
                    row.append("")

        if 'ERROR' in item:
            row.append(f"{item['ERROR']} {item.get('Details', '')}")
        else:
            row.append("")

        return row

    def _append_to_excel(self, filepath, rows, header_type="manual"):
        """
        header_type: "auto" (使用分类表头) 或 "manual" (使用数字表头)
        """
        if not os.path.exists(filepath):
            wb = openpyxl.Workbook()
            ws = wb.active

            # 根据类型决定图片列的表头
            if header_type == "auto":
                img_headers = IMG_HEADERS_AUTO
            else:
                img_headers = IMG_HEADERS_MANUAL

            headers = ["SKU"] + FIXED_TEXT_HEADERS + img_headers + ["ERROR"]
            ws.append(headers)
            wb.save(filepath)

        # 如果文件已存在，直接追加数据（此时表头由文件已有的决定，不再更改）
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        for r in rows: ws.append(r)
        wb.save(filepath)

    # ... (UI 辅助方法保持不变) ...

    def create_operation_page(self):
        op_page = QWidget()
        self.tab_widget.addTab(op_page, "操作执行")
        layout = QFormLayout(op_page)
        layout.addRow(QLabel("--- 账号档案配置 ---"))
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        account_list_widget = QWidget()
        self.account_list_layout = QVBoxLayout(account_list_widget)
        self.account_list_layout.setAlignment(Qt.AlignTop)
        scroll_area.setWidget(account_list_widget)
        scroll_area.setMaximumHeight(100)
        layout.addRow(scroll_area)
        self.account_name_input = QLineEdit()
        layout.addRow("档案名称:", self.account_name_input)
        self.username_input = QLineEdit()
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)
        layout.addRow("账号:", self.username_input)
        layout.addRow("密码:", self.password_input)
        account_btns = QHBoxLayout()
        self.save_account_button = QPushButton("保存档案")
        self.save_account_button.clicked.connect(self.save_current_account)
        self.delete_account_button = QPushButton("删除档案")
        self.delete_account_button.clicked.connect(self.delete_current_account)
        account_btns.addWidget(self.save_account_button)
        account_btns.addWidget(self.delete_account_button)
        layout.addRow(account_btns)
        layout.addRow(QLabel("--- 运行参数 ---"))
        self.url_input = QLineEdit(self.runtime_url)
        layout.addRow("目标 URL:", self.url_input)
        self.org_code_input = QLineEdit(self.runtime_org_code)
        layout.addRow("组织代码:", self.org_code_input)
        path_layout = QHBoxLayout()
        self.file_path_input = QLineEdit(self.sku_file_path)
        self.file_path_input.setReadOnly(True)
        f_btn = QPushButton("...")
        f_btn.clicked.connect(self.open_file_dialog)
        f_btn.setMaximumWidth(30)
        path_layout.addWidget(self.file_path_input)
        path_layout.addWidget(f_btn)
        layout.addRow("SKU 文件:", path_layout)
        self.start_point_combo = QComboBox()
        self.start_point_combo.addItems(["完整流程 (从登录开始)", "跳过登录/组织选择 (从导航开始)"])
        idx = self.start_point_combo.findText(self.runtime_start_point)
        if idx != -1: self.start_point_combo.setCurrentIndex(idx)
        layout.addRow("起始点:", self.start_point_combo)
        self.mode_combo = QComboBox()
        self.mode_combo.addItems(["模式一：自动导出 (按分类)", "模式二：人工筛选 (抓取所有)"])
        mode_idx = self.mode_combo.findText(self.runtime_run_mode)
        if mode_idx != -1: self.mode_combo.setCurrentIndex(mode_idx)
        layout.addRow("运行模式:", self.mode_combo)
        self.headless_checkbox = QCheckBox("后台静默运行 (无浏览器窗口)")
        self.headless_checkbox.setChecked(self.runtime_headless)
        layout.addRow("", self.headless_checkbox)
        self.execute_button = QPushButton("开始执行")
        self.execute_button.clicked.connect(self.start_automation)
        self.stop_button = QPushButton("停止运行 (完成当前批次后)")
        self.stop_button.clicked.connect(self.stop_automation)
        self.stop_button.setEnabled(False)
        self.stop_button.setStyleSheet("background-color: #f44336; color: white;")
        self.save_button = QPushButton("保存配置")
        self.save_button.clicked.connect(self.save_config)
        btns = QHBoxLayout()
        btns.addWidget(self.save_button)
        btns.addWidget(self.execute_button)
        btns.addWidget(self.stop_button)
        layout.addRow(btns)
        self.log_output = QTextEdit()
        self.log_output.setReadOnly(True)
        layout.addRow(QLabel("日志:"), self.log_output)
        self.load_account_profiles_to_ui()

    def create_config_page(self):
        config_page = QWidget()
        self.tab_widget.addTab(config_page, "元素配置")
        layout = QVBoxLayout(config_page)
        help_text = ("<b>配置指南:</b><br>"
                     "1. <b>定位字符串:</b> 支持 XPath, class, 属性, 纯文本。<br>"
                     "2. <b>位置:</b> 选择相对于定位元素的导航方向 (如父/子元素)。<br>"
                     "3. <b>Index:</b> 如果定位到多个元素，选择第几个 (默认1)。")
        layout.addWidget(QLabel(help_text))
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        config_container = QWidget()
        container_layout = QVBoxLayout(config_container)
        self.element_widgets = {}
        for module_item in self.element_config:
            module_name = module_item["module"]
            group_box = QGroupBox(module_name)
            group_layout = QFormLayout(group_box)
            for element in module_item["elements"]:
                name = element["name"]
                row_widget = QWidget()
                row_layout = QHBoxLayout(row_widget)
                row_layout.setContentsMargins(0, 0, 0, 0)
                locator_input = QLineEdit(element["locator"])
                locator_input.setPlaceholderText("定位字符串")
                position_combo = QComboBox()
                position_combo.addItems(["当前元素", "父元素", "子元素", "上一个", "下一个"])
                position_combo.setCurrentText(element.get("position", "当前元素"))
                position_combo.setFixedWidth(80)
                index_input = QLineEdit(str(element.get("index", "1")))
                index_input.setPlaceholderText("Idx")
                from PyQt5.QtGui import QIntValidator
                index_input.setValidator(QIntValidator(1, 999))
                index_input.setFixedWidth(30)
                row_layout.addWidget(locator_input, stretch=3)
                row_layout.addWidget(position_combo, stretch=1)
                row_layout.addWidget(index_input, stretch=0)
                self.element_widgets[name] = {
                    "locator": locator_input,
                    "position": position_combo,
                    "index": index_input
                }
                group_layout.addRow(name, row_widget)
            container_layout.addWidget(group_box)
        scroll_area.setWidget(config_container)
        layout.addWidget(scroll_area)
        save_config_button = QPushButton("保存元素配置")
        save_config_button.clicked.connect(self.save_config)
        layout.addWidget(save_config_button)

    def log(self, message, color="black"):
        fmt = QTextCharFormat()
        if color == "red":
            fmt.setForeground(Qt.red)
        elif color == "green":
            fmt.setForeground(Qt.green)
        elif color == "blue":
            fmt.setForeground(Qt.blue)
        else:
            fmt.setForeground(Qt.black)
        cursor = self.log_output.textCursor()
        cursor.movePosition(QTextCursor.End)
        cursor.insertText(f"{message}\n", fmt)
        self.log_output.ensureCursorVisible()

    # 其他方法保持与之前版本一致
    def load_account_profiles_to_ui(self):
        for i in reversed(range(self.account_list_layout.count())):
            widget_to_remove = self.account_list_layout.itemAt(i).widget()
            if widget_to_remove: widget_to_remove.setParent(None)
        if self.all_accounts:
            for account in self.all_accounts:
                btn = QPushButton(account["name"])
                btn.setProperty("account_name", account["name"])
                btn.clicked.connect(lambda checked, name=account["name"]: self.select_account_profile_by_name(name))
                self.account_list_layout.addWidget(btn)
            self.select_account_profile_by_name(self.runtime_selected_account_name, initial_load=True)
            self.delete_account_button.setEnabled(True)
        else:
            no_account_label = QLabel("无可用账号档案。")
            self.account_list_layout.addWidget(no_account_label)
            self.account_name_input.clear()
            self.username_input.clear()
            self.password_input.clear()
            self.delete_account_button.setEnabled(False)
            self.runtime_selected_account_name = ''

    def select_account_profile_by_name(self, name, initial_load=False):
        selected_account = None
        for account in self.all_accounts:
            if account["name"] == name: selected_account = account; break
        for i in range(self.account_list_layout.count()):
            widget = self.account_list_layout.itemAt(i).widget()
            if isinstance(widget, QPushButton):
                if widget.property("account_name") == name:
                    widget.setStyleSheet("background-color: #AECBFA; font-weight: bold;")
                else:
                    widget.setStyleSheet("")
        if selected_account:
            self.account_name_input.setText(selected_account.get("name", ""))
            self.username_input.setText(selected_account.get("username", ""))
            self.password_input.setText(selected_account.get("password", ""))
            self.delete_account_button.setEnabled(True)
            self.runtime_selected_account_name = name
            if not initial_load: self.save_config()
        else:
            self.account_name_input.clear()
            self.username_input.clear()
            self.password_input.clear()
            self.delete_account_button.setEnabled(False)
            self.runtime_selected_account_name = ''

    def save_current_account(self):
        name = self.account_name_input.text().strip()
        username = self.username_input.text().strip()
        password = self.password_input.text().strip()
        if not name or not username or not password:
            QMessageBox.warning(self, "保存失败", "所有字段都不能为空。")
            return
        existing_index = -1
        for i, account in enumerate(self.all_accounts):
            if account["name"] == name: existing_index = i; break
        new_account = {"name": name, "username": username, "password": password}
        if existing_index != -1:
            self.all_accounts[existing_index] = new_account
            self.log(f"账号档案 '{name}' 已更新。", "blue")
        else:
            self.all_accounts.append(new_account)
            self.log(f"账号档案 '{name}' 已新增。", "blue")
        self.save_config()
        self.load_account_profiles_to_ui()
        self.select_account_profile_by_name(name)

    def delete_current_account(self):
        current_name = self.account_name_input.text().strip()
        if not current_name: return
        if QMessageBox.question(self, '确认删除', f"确定要删除 '{current_name}' 吗?",
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            self.all_accounts = [acc for acc in self.all_accounts if acc["name"] != current_name]
            self.log(f"已删除。", "blue")
            self.runtime_selected_account_name = ''
            self.save_config()
            self.load_account_profiles_to_ui()
            if self.all_accounts:
                self.select_account_profile_by_name(self.all_accounts[0]['name'])
            else:
                self.select_account_profile_by_name('')

    def open_file_dialog(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "选择文件", "", "Excel/CSV (*.xlsx *.csv);;All (*)")
        if file_path:
            self.file_path_input.setText(file_path)
            self.save_config()

    def get_table_data(self):
        updated_config = []
        for module_item in self.element_config:
            new_module = {"module": module_item["module"], "elements": []}
            for element in module_item["elements"]:
                name = element["name"]
                widgets = self.element_widgets.get(name)
                if widgets:
                    new_element = {
                        "name": name,
                        "locator": widgets["locator"].text(),
                        "position": widgets["position"].currentText(),
                        "index": widgets["index"].text() or "1"
                    }
                    new_module["elements"].append(new_element)
                else:
                    new_module["elements"].append(element)
            updated_config.append(new_module)
        return updated_config


# --- 补充 ClickableImageLabel 和 SKUResultWidget 类 (与上一版本一致) ---
class ClickableImageLabel(QLabel):
    def __init__(self, img_url, parent_widget):
        super().__init__()
        self.img_url = img_url
        self.parent_widget = parent_widget
        self.selection_number = 0
        self.setFixedSize(100, 100)
        self.setScaledContents(True)
        self.setCursor(Qt.PointingHandCursor)
        self.default_style = "border: 1px solid #ccc; background-color: #f0f0f0;"
        self.setStyleSheet(self.default_style)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.parent_widget.handle_image_click(self.img_url)

    def set_selection_number(self, number):
        self.selection_number = number
        self.repaint()

    def paintEvent(self, event):
        super().paintEvent(event)
        if self.selection_number > 0:
            painter = QPainter(self)
            painter.setRenderHint(QPainter.Antialiasing)
            pen = QPen(Qt.red);
            pen.setWidth(4);
            painter.setPen(pen)
            painter.drawRect(self.rect())
            badge_size = 24;
            rect = QRect(self.width() - badge_size, 0, badge_size, badge_size)
            painter.setBrush(QBrush(Qt.red));
            painter.setPen(Qt.NoPen);
            painter.drawRect(rect)
            painter.setPen(Qt.white);
            painter.setFont(QFont("Arial", 12, QFont.Bold))
            painter.drawText(rect, Qt.AlignCenter, str(self.selection_number))


class SKUResultWidget(QGroupBox):
    def __init__(self, sku_data, image_session, parent=None):
        super().__init__(parent)
        self.sku_data = sku_data
        self.image_session = image_session
        self.setTitle(f"SKU: {sku_data['SKU']}")
        self.selected_images_list = []
        self.image_label_map = {}
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        if 'ERROR' in self.sku_data:
            layout.addWidget(QLabel(f"❌ {self.sku_data['ERROR']}", styleSheet="color: red;"))

        images = self.sku_data.get('all_images', [])
        if images:
            area = QWidget();
            grid = QGridLayout(area)
            for i, url in enumerate(images):
                lbl = ClickableImageLabel(url, self)
                lbl.setToolTip(url)
                self.image_label_map[url] = lbl
                if i < 15:  # 预览限制
                    try:
                        r = self.image_session.get(url, timeout=1)
                        if r.status_code == 200:
                            p = QPixmap();
                            p.loadFromData(r.content);
                            lbl.setPixmap(p)
                        else:
                            lbl.setText("X")
                    except:
                        lbl.setText("Err")
                else:
                    lbl.setText(f"Pic {i + 1}")
                grid.addWidget(lbl, i // 8, i % 8)
            layout.addWidget(area)
        else:
            layout.addWidget(QLabel("无图片"))
        self.setLayout(layout)

    def handle_image_click(self, url):
        if url in self.selected_images_list:
            self.selected_images_list.remove(url)
        else:
            if len(self.selected_images_list) < 9: self.selected_images_list.append(url)
        self.refresh()

    def refresh(self):
        for i, url in enumerate(self.selected_images_list):
            if url in self.image_label_map: self.image_label_map[url].set_selection_number(i + 1)
        for url, lbl in self.image_label_map.items():
            if url not in self.selected_images_list: lbl.set_selection_number(0)

    def get_selected_images(self):
        return self.selected_images_list


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = AutomationToolUI()
    window.show()
    sys.exit(app.exec_())