import sys
import os
import json
import time
import pandas as pd
import openpyxl
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QFormLayout, QLineEdit, QPushButton,
                             QLabel, QGroupBox, QRadioButton, QButtonGroup,
                             QFileDialog, QTextEdit, QProgressBar, QMessageBox, QSplitter)
from PyQt5.QtCore import QThread, pyqtSignal, Qt
from openai import OpenAI

# ================= 配置区域 =================
# 【重要】请在这里硬编码你的 API Key
DEFAULT_API_KEY = "sk-a3cb070863f745c1bbdbe63083d12757"
DEEPSEEK_BASE_URL = "https://api.deepseek.com"
USER_CONFIG_FILE = "user_settings.json"

# --- 【修改点 1】 更新 System Prompt，整合所有新要求 ---
DEFAULT_SYSTEM_PROMPT = """你是一位亚马逊 (Amazon) Listing 资深文案专家。
请根据提供的产品信息，严格按照以下规则编写 SEO 优化内容。

【1. 标题要求 (Title)】
- **长度**：控制在 150 字符以内。
- **格式**：必须严格遵守 "[核心关键词], [参数 + 特征 + 使用场景 + 目标人群]"。
- **标点**：全标题只允许出现**一个逗号**，且必须紧跟在核心关键词之后。
- **禁忌**：禁止使用品牌名、宣传语（如 Best Seller, Hot）或特殊符号。

【2. 关键词要求 (Search Terms)】
- **长度**：总长度控制在 150 - 200 字符之间。
- **格式**：仅使用**空格**分隔单词，严禁使用逗号(,)、分号(;)或横杠(-)。
- **内容**：不得重复单词，包含尽可能多的长尾词和相关搜索词。

【3. 五点描述要求 (Bullet Points)】
- **格式**：每点开头用全大写短语概括 (如 "HIGH PERFORMANCE:")。
- **内容分配**：
  - Point 1 (核心): 型号、技术参数、核心功能。
  - Point 2 (材质): 材质及其耐用/特性。
  - Point 3 (尺寸): 必须同时包含 **英寸(inch)** 和 **毫米(mm)** 单位。
  - Point 4 (场景): 适用场景和具体目标用户群体。
  - Point 5 (附加): 附加优势、易用性、安装建议或保养建议。

【4. 描述要求 (Description)】
- 通顺的营销短文，使用 <br> 进行段落换行，突出产品解决的痛点。

【5. 输出格式】
- 必须是严格的 JSON 格式，不要包含 Markdown 标记（如 ```json）。
- 即使原数据缺失，也要根据常识生成合理内容。

【JSON 模板】
{
    "title": "Core Keyword, features usage for target audience...",
    "keywords": "keyword1 keyword2 keyword3 ...",
    "bullet1": "KEY FEATURE: ...",
    "bullet2": "MATERIAL: ...",
    "bullet3": "SIZE SPECS: ... (xx inch / xx mm)",
    "bullet4": "USAGE: ...",
    "bullet5": "ADDITIONAL: ...",
    "description": "..."
}
"""

# 列名映射配置
COLUMN_MAPPING = {
    "title": ["Item Title", "Title", "标题", "产品标题"],
    "keywords": ["Key Words", "Keywords", "关键词", "搜索词"],
    "bullet1": ["五点1", "Bullet Point 1", "BulletPoint1", "卖点1"],
    "bullet2": ["五点2", "Bullet Point 2", "BulletPoint2", "卖点2"],
    "bullet3": ["五点3", "Bullet Point 3", "BulletPoint3", "卖点3"],
    "bullet4": ["五点4", "Bullet Point 4", "BulletPoint4", "卖点4"],
    "bullet5": ["五点5", "Bullet Point 5", "BulletPoint5", "卖点5"],
    "description": ["Description", "Product Description", "描述", "产品描述"],
    # 辅助信息
    "category": ["商品目录", "Category"],
    "size": ["尺寸(cm)", "Size"],
    "weight": ["重量(g)", "Weight"]
}


# ===========================================

class WorkerThread(QThread):
    progress_signal = pyqtSignal(int, int)  # 当前, 总数
    log_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(str)
    error_signal = pyqtSignal(str)
    status_signal = pyqtSignal(bool)  # True=Running, False=Stopped

    def __init__(self, api_key, model, temp, file_path, system_prompt):
        super().__init__()
        self.api_key = api_key
        self.model = model
        self.temp = temp
        self.file_path = file_path
        self.system_prompt = system_prompt
        self.is_running = True

    def run(self):
        try:
            self.status_signal.emit(True)
            client = OpenAI(api_key=self.api_key, base_url=DEEPSEEK_BASE_URL)

            # 1. 读取源数据
            self.log_signal.emit(f"正在读取源文件: {os.path.basename(self.file_path)} ...")
            try:
                if self.file_path.endswith('.csv'):
                    df_source = pd.read_csv(self.file_path)
                else:
                    df_source = pd.read_excel(self.file_path, engine='openpyxl')
            except Exception as e:
                error_msg = str(e)
                if "No such keys" in error_msg or "BadZipFile" in error_msg:
                    raise ValueError(
                        f"文件已损坏，无法读取！\n原因：该 Excel 文件可能未正常保存或被强制中断。\n建议：请删除该文件，重新运行采集工具生成，或尝试用 Excel 打开并另存为修复。")
                else:
                    raise e

            total_rows = len(df_source)
            self.log_signal.emit(f"读取成功，共 {total_rows} 行数据。")

            # 2. 准备 Output Excel
            try:
                wb = openpyxl.load_workbook(self.file_path)
            except Exception as e:
                raise ValueError(f"无法加载 Excel (可能文件损坏): {e}")

            target_sheet_name = "AI_Optimized"

            if target_sheet_name in wb.sheetnames:
                ws = wb[target_sheet_name]
                self.log_signal.emit(f"发现已有优化表 '{target_sheet_name}'，将进行增量更新。")
            else:
                ws = wb.create_sheet(target_sheet_name)
                # --- 【修改点 2】 表头增加 AI_Title ---
                headers = ["SKU", "AI_Title", "AI_Keywords", "AI_Bullet1", "AI_Bullet2", "AI_Bullet3", "AI_Bullet4",
                           "AI_Bullet5", "AI_Description"]
                ws.append(headers)
                try:
                    wb.save(self.file_path)
                except PermissionError:
                    raise ValueError("文件被占用！请先关闭 Excel 文件。")
                self.log_signal.emit(f"新建优化表 '{target_sheet_name}'。")

            # 3. 获取已处理的 SKU 列表
            processed_skus = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    sku_val = str(row[0]).strip()
                    # 简单检查是否已有内容 (检查第3列 keywords 是否有值)
                    has_content = False
                    if len(row) > 2 and row[2] and str(row[2]).strip():
                        has_content = True

                    if has_content:
                        processed_skus.add(sku_val)

            if processed_skus:
                self.log_signal.emit(f"检测到 {len(processed_skus)} 个已完成任务，将自动跳过。")

            # 4. 映射列名
            col_map = {}
            for key, possible_names in COLUMN_MAPPING.items():
                for col in df_source.columns:
                    if any(name in str(col).strip() for name in possible_names):
                        col_map[key] = col
                        break

            # 5. 循环处理
            processed_count = 0

            for index, row in df_source.iterrows():
                if not self.is_running:
                    self.log_signal.emit("⚠️ 任务已由用户手动停止。")
                    break

                sku = str(row[df_source.columns[0]]).strip()

                if sku in processed_skus:
                    self.progress_signal.emit(index + 1, total_rows)
                    continue

                self.log_signal.emit(f"正在优化 ({index + 1}/{total_rows}): SKU {sku} ...")

                product_info = {
                    "sku": sku,
                    "title": str(row[col_map['title']]) if 'title' in col_map else "",
                    "category": str(row[col_map['category']]) if 'category' in col_map else "",
                    "size": str(row[col_map['size']]) if 'size' in col_map else "",
                    "weight": str(row[col_map['weight']]) if 'weight' in col_map else "",
                    "keywords": str(row[col_map['keywords']]) if 'keywords' in col_map else "",
                    "bullets": [str(row[col_map[f'bullet{i}']]) if f'bullet{i}' in col_map else "" for i in
                                range(1, 6)],
                    "description": str(row[col_map['description']]) if 'description' in col_map else ""
                }

                # 即使标题为空，只要有其他信息也可以尝试生成，但这里保留原逻辑跳过
                if not product_info['title'] or product_info['title'] == 'nan':
                    self.log_signal.emit("  -> 跳过: 原标题为空，无法参考")
                    continue

                result = self.call_ai(client, product_info)

                if result:
                    # --- 【修改点 3】 保存 AI 生成的标题 (如果生成失败则用原标题兜底) ---
                    ai_title = result.get("title", product_info['title'])

                    new_row = [
                        sku,
                        ai_title,
                        result.get("keywords", ""),
                        result.get("bullet1", ""),
                        result.get("bullet2", ""),
                        result.get("bullet3", ""),
                        result.get("bullet4", ""),
                        result.get("bullet5", ""),
                        result.get("description", "")
                    ]

                    ws.append(new_row)
                    try:
                        wb.save(self.file_path)
                        self.log_signal.emit("  -> ✅ 已保存")
                        processed_count += 1
                    except PermissionError:
                        self.log_signal.emit("  -> ❌ 保存失败: 请先关闭 Excel 文件！")
                        self.is_running = False
                        break
                else:
                    self.log_signal.emit("  -> ❌ 优化失败")

                self.progress_signal.emit(index + 1, total_rows)
                time.sleep(0.5)

            if self.is_running:
                self.finished_signal.emit(f"所有任务完成！共新增处理 {processed_count} 条数据。")
            else:
                self.finished_signal.emit(f"任务已暂停/停止。本次共处理 {processed_count} 条数据。")

            self.status_signal.emit(False)

        except Exception as e:
            self.error_signal.emit(str(e))
            self.status_signal.emit(False)

    def call_ai(self, client, info):
        # 提示词中加入更详细的尺寸单位提醒
        user_prompt = f"""
请基于以下产品原数据进行优化：
SKU: {info['sku']}
产品目录: {info['category']}
尺寸(原数据): {info['size']} (注意：五点描述第3点必须包含 Inch 和 mm 双单位)
重量: {info['weight']}
原标题: {info['title']}
原关键词: {info['keywords']}
原五点: {info['bullets']}
原描述: {info['description']}
"""
        response_format = {"type": "json_object"} if self.model == "deepseek-chat" else None

        try:
            response = client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": self.system_prompt},
                    {"role": "user", "content": user_prompt}
                ],
                temperature=self.temp,
                response_format=response_format,
                max_tokens=4096
            )
            content = response.choices[0].message.content.replace("```json", "").replace("```", "").strip()
            return json.loads(content)
        except Exception as e:
            print(f"AI Error: {e}")
            return None

    def stop(self):
        self.is_running = False


class AICopywriterUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("亚马逊文案 AI 优化工具 (DeepSeek内核)")
        self.setGeometry(300, 100, 1000, 800)

        main_widget = QWidget()
        self.setCentralWidget(main_widget)

        main_layout = QHBoxLayout(main_widget)

        # --- 左侧面板 ---
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)

        # 1. 参数设置
        param_group = QGroupBox("模型参数 (API Key 已内置)")
        param_layout = QFormLayout(param_group)

        self.model_group = QButtonGroup(self)
        self.rb_model_v3 = QRadioButton("DeepSeek-V3 (Chat)")
        self.rb_model_r1 = QRadioButton("DeepSeek-R1 (Reasoner)")
        self.rb_model_v3.setChecked(True)
        self.model_group.addButton(self.rb_model_v3)
        self.model_group.addButton(self.rb_model_r1)

        model_layout = QHBoxLayout()
        model_layout.addWidget(self.rb_model_v3)
        model_layout.addWidget(self.rb_model_r1)
        param_layout.addRow("AI 模型:", model_layout)

        self.temp_group = QButtonGroup(self)
        self.rb_temp_1_0 = QRadioButton("1.0 (标准)")
        self.rb_temp_1_3 = QRadioButton("1.3 (创意) [推荐]")
        self.rb_temp_1_3.setChecked(True)
        self.temp_group.addButton(self.rb_temp_1_0)
        self.temp_group.addButton(self.rb_temp_1_3)

        temp_layout = QHBoxLayout()
        temp_layout.addWidget(self.rb_temp_1_0)
        temp_layout.addWidget(self.rb_temp_1_3)
        param_layout.addRow("温度:", temp_layout)
        left_layout.addWidget(param_group)

        # 2. 文件操作
        file_group = QGroupBox("文件操作")
        file_layout = QFormLayout(file_group)
        self.file_path_input = QLineEdit()
        self.file_path_input.setReadOnly(True)
        btn_select = QPushButton("选择 Excel")
        btn_select.clicked.connect(self.select_file)
        file_layout.addRow(self.file_path_input)
        file_layout.addRow(btn_select)
        left_layout.addWidget(file_group)

        # 3. 控制按钮区
        btn_layout = QHBoxLayout()

        # 保存配置按钮
        self.btn_save_config = QPushButton("保存配置")
        self.btn_save_config.setFixedHeight(40)
        self.btn_save_config.clicked.connect(self.manual_save_settings)

        self.btn_start = QPushButton("开始优化")
        self.btn_start.setFixedHeight(40)
        self.btn_start.setStyleSheet("font-weight: bold; background-color: #0078D7; color: white;")
        self.btn_start.clicked.connect(self.toggle_optimization)

        btn_layout.addWidget(self.btn_save_config)
        btn_layout.addWidget(self.btn_start)
        left_layout.addLayout(btn_layout)

        # 4. 进度条
        self.progress_bar = QProgressBar()
        left_layout.addWidget(self.progress_bar)

        left_layout.addStretch()

        # --- 右侧面板 ---
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)

        prompt_group = QGroupBox("AI 角色指令 (System Prompt) - 可编辑")
        prompt_layout = QVBoxLayout(prompt_group)
        self.prompt_edit = QTextEdit()
        self.prompt_edit.setPlainText(DEFAULT_SYSTEM_PROMPT)
        self.prompt_edit.setStyleSheet("font-family: Consolas, monospace; font-size: 12px;")
        prompt_layout.addWidget(self.prompt_edit)
        right_layout.addWidget(prompt_group, stretch=2)

        log_group = QGroupBox("执行日志")
        log_layout = QVBoxLayout(log_group)
        self.log_output = QTextEdit()
        self.log_output.setReadOnly(True)
        log_layout.addWidget(self.log_output)
        right_layout.addWidget(log_group, stretch=1)

        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)
        splitter.setSizes([300, 600])

        main_layout.addWidget(splitter)

        self.worker = None
        self.load_settings()

    def load_settings(self):
        """加载本地配置文件"""
        if os.path.exists(USER_CONFIG_FILE):
            try:
                with open(USER_CONFIG_FILE, 'r', encoding='utf-8') as f:
                    config = json.load(f)

                if config.get("model") == "deepseek-reasoner":
                    self.rb_model_r1.setChecked(True)
                else:
                    self.rb_model_v3.setChecked(True)

                if config.get("temperature") == 1.0:
                    self.rb_temp_1_0.setChecked(True)
                else:
                    self.rb_temp_1_3.setChecked(True)

                saved_prompt = config.get("system_prompt", "")
                if saved_prompt:
                    self.prompt_edit.setPlainText(saved_prompt)
                else:
                    self.prompt_edit.setPlainText(DEFAULT_SYSTEM_PROMPT)
                self.log("✅ 已加载本地配置 (模型/Prompt)。")
            except Exception as e:
                self.log(f"⚠️ 配置文件加载失败: {e}")
        else:
            self.save_settings()
            self.log("🆕 首次运行，已创建默认配置文件。")

    def save_settings(self):
        """保存配置到本地"""
        config = {
            "model": "deepseek-chat" if self.rb_model_v3.isChecked() else "deepseek-reasoner",
            "temperature": 1.3 if self.rb_temp_1_3.isChecked() else 1.0,
            "system_prompt": self.prompt_edit.toPlainText()
        }
        try:
            with open(USER_CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=4)
            return True
        except Exception as e:
            self.log(f"❌ 配置保存失败: {e}")
            return False

    def manual_save_settings(self):
        if self.save_settings():
            QMessageBox.information(self, "保存成功", "参数配置已保存 (user_settings.json)")
            self.log("配置已手动保存。")

    def select_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择表格", "", "Excel Files (*.xlsx)")
        if path:
            self.file_path_input.setText(path)

    def log(self, msg):
        self.log_output.append(msg)
        cursor = self.log_output.textCursor()
        cursor.movePosition(cursor.End)
        self.log_output.setTextCursor(cursor)

    def toggle_optimization(self):
        if self.worker and self.worker.isRunning():
            self.log("正在请求停止... 请等待当前 SKU 处理完成...")
            self.btn_start.setText("正在停止...")
            self.btn_start.setEnabled(False)
            self.worker.stop()
            return

        input_path = self.file_path_input.text().strip()
        api_key = DEFAULT_API_KEY

        if "sk-" not in api_key:
            QMessageBox.critical(self, "配置错误", "代码中的 DEFAULT_API_KEY 无效，请检查代码！")
            return

        if not input_path or not os.path.exists(input_path):
            QMessageBox.warning(self, "错误", "请选择有效的 Excel 文件")
            return

        try:
            with open(input_path, "a"):
                pass
        except PermissionError:
            QMessageBox.critical(self, "无法访问", "文件被占用！\n请先关闭 Excel 文件再运行。")
            return

        self.save_settings()

        model = "deepseek-chat" if self.rb_model_v3.isChecked() else "deepseek-reasoner"
        temp = 1.3 if self.rb_temp_1_3.isChecked() else 1.0
        system_prompt = self.prompt_edit.toPlainText()

        self.log(f"--- 任务启动 ---")
        self.log(f"模型: {model}, 温度: {temp}")
        self.btn_start.setText("停止优化 (保存并释放表格)")
        self.btn_start.setStyleSheet("font-size: 16px; font-weight: bold; background-color: #D9534F; color: white;")
        self.progress_bar.setValue(0)

        self.worker = WorkerThread(api_key, model, temp, input_path, system_prompt)
        self.worker.log_signal.connect(self.log)
        self.worker.progress_signal.connect(self.update_progress)
        self.worker.finished_signal.connect(self.task_finished)
        self.worker.error_signal.connect(self.task_error)
        self.worker.status_signal.connect(self.update_ui_state)
        self.worker.start()

    def update_progress(self, current, total):
        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(current)

    def task_finished(self, msg):
        QMessageBox.information(self, "状态", msg)
        self.log(msg)

    def task_error(self, msg):
        QMessageBox.critical(self, "出错", f"运行中发生错误:\n{msg}")
        self.log(f"❌ 错误: {msg}")

    def update_ui_state(self, is_running):
        if not is_running:
            self.btn_start.setText("开始优化")
            self.btn_start.setEnabled(True)
            self.btn_start.setStyleSheet("font-size: 16px; font-weight: bold; background-color: #0078D7; color: white;")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = AICopywriterUI()
    window.show()
    sys.exit(app.exec_())